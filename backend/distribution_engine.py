"""
Distribution Engine
====================
Executes a distribution rule: loads the report template, fetches tag data,
generates HTML/PDF, and delivers via email and/or disk.
"""

import os
import re
import json
import logging
import smtplib
import base64
from datetime import datetime, timedelta
from email.message import EmailMessage
from html import escape as html_escape
from contextlib import closing

from psycopg2.extras import RealDictCursor
from io import BytesIO

logger = logging.getLogger(__name__)


# ── Chart image generation (matplotlib) ──────────────────────────────────────

def _render_chart_image_base64(chart_type, series_data, label='', width_px=500, height_px=200):
    """Render a chart as a base64 PNG data URI using matplotlib.
    series_data: list of (label, value) tuples.
    Returns data:image/png;base64,... string, or None on failure.
    """
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt

        fig_w = max(3, width_px / 100)
        fig_h = max(1.5, height_px / 100)
        fig, ax = plt.subplots(figsize=(fig_w, fig_h), dpi=100)

        labels = [s[0] for s in series_data]
        values = [s[1] for s in series_data]
        colors = ['#3b82f6', '#10b981', '#f59e0b', '#ef4444', '#8b5cf6', '#06b6d4', '#ec4899', '#6366f1']

        if chart_type == 'piechart' and any(v > 0 for v in values):
            safe_vals = [max(v, 0) for v in values]
            ax.pie(safe_vals, labels=labels, autopct='%1.0f%%', startangle=90,
                   colors=colors[:len(values)], textprops={'fontsize': 8})
        else:
            bars = ax.bar(labels, values, color=colors[:len(values)], width=0.6)
            ax.set_ylabel('')
            ax.tick_params(axis='x', labelsize=7, rotation=30)
            ax.tick_params(axis='y', labelsize=7)
            for bar, val in zip(bars, values):
                ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height(),
                        f'{val:,.1f}', ha='center', va='bottom', fontsize=7)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)

        if label and chart_type != 'piechart':
            ax.set_title(label, fontsize=9, fontweight='bold', pad=6)

        plt.tight_layout(pad=0.5)
        buf = BytesIO()
        fig.savefig(buf, format='png', bbox_inches='tight', dpi=100)
        plt.close(fig)
        buf.seek(0)
        b64 = base64.b64encode(buf.read()).decode('ascii')
        return f'data:image/png;base64,{b64}'
    except Exception as e:
        logger.debug("Chart rendering failed: %s", e)
        return None

# Default save directory (used when no path is provided)
DEFAULT_SAVE_DIR = os.getenv(
    'REPORT_SAVE_ROOT',
    os.path.join(os.path.abspath(os.path.dirname(__file__)), 'reports')
)


def _get_db_connection():
    """Get database connection function, avoiding circular imports."""
    import sys
    if 'app' in sys.modules:
        fn = getattr(sys.modules['app'], 'get_db_connection', None)
        if fn:
            return fn
    raise RuntimeError("Could not get database connection function")


# ── Tag extraction from layout_config ────────────────────────────────────────

_FORMULA_AGG_PREFIX_RE = re.compile(r'^(first|last|delta|avg|min|max|sum|count)::(.+)$', re.I)


def _parse_formula_tag_tokens(formula):
    """Yield (base_tag_name, explicit_agg_or_None) for each {…} token, excluding {col:…}."""
    for m in re.finditer(r'\{([^}]+)\}', formula or ''):
        inner = m.group(1)
        if inner.startswith('col:'):
            continue
        mm = _FORMULA_AGG_PREFIX_RE.match(inner)
        if mm:
            yield mm.group(2), mm.group(1).lower()
        else:
            yield inner, None


def _parse_formula_tags(formula):
    """Extract base tag names from formula strings (paginated uses {Tag} or {first::Tag})."""
    tags = set()
    tags.update(re.findall(r'\[([^\]]+)\]', formula or ''))
    for base, _explicit in _parse_formula_tag_tokens(formula):
        tags.add(base)
    return tags


def _extract_datasource_tags(ds):
    """Extract tag names from a single dataSource dict."""
    tags = set()
    ds_type = ds.get('type', 'tag')
    if ds_type == 'tag' and ds.get('tagName'):
        tags.add(ds['tagName'])
    elif ds_type == 'group' and ds.get('groupTags'):
        tags.update(ds['groupTags'])
    elif ds_type == 'formula' and ds.get('formula'):
        tags.update(_parse_formula_tags(ds['formula']))
    return tags


def _chart_series_list_for_export(config):
    """Series list for chart/barchart/piechart — matches ChartWidget (series or legacy tags[])."""
    if not isinstance(config, dict):
        return []
    raw = config.get('series')
    if isinstance(raw, list) and len(raw) > 0:
        return raw
    built = []
    for t in config.get('tags') or []:
        if isinstance(t, dict):
            tn = t.get('tagName', '')
            if tn:
                built.append({
                    'label': t.get('displayName') or t.get('label') or tn,
                    'dataSource': {'type': 'tag', 'tagName': tn},
                })
        elif isinstance(t, str) and t.strip():
            built.append({'label': t, 'dataSource': {'type': 'tag', 'tagName': t.strip()}})
    return built


def _flatten_tabcontainer_widgets(tc_widget):
    """Ordered list of inner widgets from all tabs of a tabcontainer, expanding nested tabcontainers."""
    out = []
    cfg = (tc_widget or {}).get('config') or {}
    for tab in cfg.get('tabs') or []:
        for w in tab.get('widgets') or []:
            if not w:
                continue
            if w.get('type') == 'tabcontainer':
                out.extend(_flatten_tabcontainer_widgets(w))
            else:
                out.append(w)
    return out


def _flatten_dashboard_widget_list(widgets):
    """Top-level dashboard widgets with tabcontainer nodes expanded."""
    out = []
    for w in widgets or []:
        if not w:
            continue
        if w.get('type') == 'tabcontainer':
            out.extend(_flatten_tabcontainer_widgets(w))
        else:
            out.append(w)
    return out


def _dashboard_sections_for_distribution(layout_config):
    """Sections of (optional heading, widgets) for PDF/XLSX/email.

    When dashboard multi-tabs are enabled (Report Builder), widgets are stored per tab
    under layout_config.dashboardTabs and the root ``widgets`` array is empty — the
    distribution engine must read each tab's widget list.

    Returns:
        list[tuple[str | None, list]] — non-empty widget lists only (except fallback).
    """
    lc = layout_config if isinstance(layout_config, dict) else {}
    dt = lc.get('dashboardTabs') or {}
    tabs = dt.get('tabs') if isinstance(dt.get('tabs'), list) else []
    if dt.get('enabled') and tabs:
        sections = []
        for tab in tabs:
            label = tab.get('label')
            if label is not None:
                label = str(label).strip() or None
            flat = _flatten_dashboard_widget_list(tab.get('widgets') or [])
            if flat:
                sections.append((label, flat))
        if sections:
            return sections
    root = lc.get('widgets') or []
    return [(None, _flatten_dashboard_widget_list(root))]


def _update_tags_from_dashboard_widget(widget, tags):
    """Merge tag names referenced by one dashboard widget into ``tags``."""
    config = widget.get('config', {}) or {}
    widget_type = widget.get('type', '')

    ds = config.get('dataSource', {}) or {}
    tags.update(_extract_datasource_tags(ds))

    if widget_type in ('silo', 'hopper'):
        for key in ('capacityTag', 'tonsTag'):
            if config.get(key):
                tags.add(config[key])

    for series in _chart_series_list_for_export(config):
        s_ds = series.get('dataSource', {}) or {}
        tags.update(_extract_datasource_tags(s_ds))
        if series.get('tagName'):
            tags.add(series['tagName'])

    for col in (config.get('tableColumns') or config.get('columns') or []):
        col_type = col.get('sourceType', 'tag')
        if col_type == 'tag' and col.get('tagName'):
            tags.add(col['tagName'])
        elif col_type == 'group' and col.get('groupTags'):
            tags.update(col['groupTags'])
        elif col_type == 'formula' and col.get('formula'):
            tags.update(_parse_formula_tags(col['formula']))

    # datapanel fields
    for field in config.get('fields', []):
        if field.get('tagName'):
            tags.add(field['tagName'])
        if field.get('formula'):
            tags.update(_parse_formula_tags(field['formula']))

    # statusbar tags
    for st in config.get('tags', []):
        if isinstance(st, dict) and st.get('tagName'):
            tags.add(st['tagName'])


def _extract_paginated_tags(sections):
    """Extract tag names from paginatedSections (Table Report format)."""
    tags = set()
    for s in (sections or []):
        s_type = s.get('type', '')

        if s_type == 'header':
            if s.get('statusTagName'):
                tags.add(s['statusTagName'])
            if s.get('statusFormula'):
                tags.update(_parse_formula_tags(s['statusFormula']))
            if s.get('statusSourceType') == 'group':
                for t in (s.get('statusGroupTags') or []):
                    if t:
                        tags.add(t)

        elif s_type == 'kpi-row':
            for k in (s.get('kpis') or []):
                if k.get('tagName'):
                    tags.add(k['tagName'])
                if k.get('formula'):
                    tags.update(_parse_formula_tags(k['formula']))
                if k.get('sourceType') == 'group':
                    for t in (k.get('groupTags') or []):
                        if t:
                            tags.add(t)

        elif s_type == 'table':
            for row in (s.get('rows') or []):
                for cell in (row.get('cells') or []):
                    if cell.get('tagName'):
                        tags.add(cell['tagName'])
                    if cell.get('formula'):
                        tags.update(_parse_formula_tags(cell['formula']))
                    if cell.get('sourceType') == 'group':
                        for t in (cell.get('groupTags') or []):
                            if t:
                                tags.add(t)
            if s.get('summaryFormula'):
                tags.update(_parse_formula_tags(s['summaryFormula']))
            for col in (s.get('columns') or []):
                sf = col.get('summary', {}).get('formula')
                if sf:
                    tags.update(_parse_formula_tags(sf))

    return tags


def extract_all_tags(layout_config):
    """Extract every tag name referenced in a report layout_config."""
    tags = set()

    # Dashboard widgets (root grid, dashboardTabs.*.widgets, nested tabcontainers)
    for _heading, widget_list in _dashboard_sections_for_distribution(layout_config):
        for widget in widget_list:
            _update_tags_from_dashboard_widget(widget, tags)

    # Paginated (Table Report) sections
    tags.update(_extract_paginated_tags(layout_config.get('paginatedSections', [])))

    return tags


def _collect_aggregation_groups(layout_config):
    """Collect per-tag aggregation types from paginated sections.
    Returns { aggregation_type: set(tag_names) } so we can fetch each group separately.
    Tags with no explicit aggregation default to 'last'.
    """
    agg_groups = {}  # { 'last': set(), 'first': set(), 'delta': set(), ... }

    def add_tag(tag_name, aggregation):
        agg = aggregation or 'last'
        if agg not in agg_groups:
            agg_groups[agg] = set()
        agg_groups[agg].add(tag_name)

    for s in (layout_config.get('paginatedSections') or []):
        s_type = s.get('type', '')

        if s_type == 'kpi-row':
            for k in (s.get('kpis') or []):
                if k.get('tagName'):
                    add_tag(k['tagName'], k.get('aggregation'))
                if k.get('formula'):
                    for base, explicit in _parse_formula_tag_tokens(k['formula']):
                        add_tag(base, explicit or k.get('aggregation'))

        elif s_type == 'table':
            for row in (s.get('rows') or []):
                for cell in (row.get('cells') or []):
                    src = cell.get('sourceType', 'static')
                    if src == 'tag' and cell.get('tagName'):
                        add_tag(cell['tagName'], cell.get('aggregation'))
                    elif src == 'formula' and cell.get('formula'):
                        for base, explicit in _parse_formula_tag_tokens(cell['formula']):
                            add_tag(base, explicit or cell.get('aggregation'))

    return agg_groups


def _fetch_tag_data_multi_agg(layout_config, tag_names, from_dt, to_dt):
    """Fetch tag data with per-cell aggregation support.
    For non-'last' aggregations, keys are namespaced as 'agg::tagName'.
    Returns a single merged dict.
    """
    agg_groups = _collect_aggregation_groups(layout_config)

    # Always fetch 'last' for all tags (default)
    tag_data = _fetch_tag_data(tag_names, from_dt, to_dt, aggregation='last')

    # Fetch additional aggregation groups and namespace the keys
    for agg, agg_tags in agg_groups.items():
        if agg == 'last':
            continue  # already fetched
        agg_result = _fetch_tag_data(agg_tags, from_dt, to_dt, aggregation=agg)
        for tag_name, value in agg_result.items():
            tag_data[f'{agg}::{tag_name}'] = value

    return tag_data


# ── Time range helper ────────────────────────────────────────────────────────

def _time_range_for_schedule(schedule_type):
    """Return (from_dt, to_dt) based on schedule frequency."""
    now = datetime.now()
    if schedule_type == 'daily':
        return now - timedelta(hours=24), now
    elif schedule_type == 'weekly':
        return now - timedelta(days=7), now
    else:  # monthly
        return now - timedelta(days=30), now


# ── Fetch tag data ───────────────────────────────────────────────────────────

def _fetch_tag_data(tag_names, from_dt, to_dt, aggregation='last'):
    """Fetch aggregated tag values from historian — same SQL as historian_bp /by-tags."""
    if not tag_names:
        return {}

    get_conn = _get_db_connection()
    with closing(get_conn()) as conn:
        actual_conn = conn._conn if hasattr(conn, '_conn') else conn
        cur = actual_conn.cursor(cursor_factory=RealDictCursor)

        cur.execute(
            "SELECT id, tag_name FROM tags WHERE tag_name = ANY(%s) AND is_active = true",
            (list(tag_names),)
        )
        tag_map = {row['tag_name']: row['id'] for row in cur.fetchall()}
        if not tag_map:
            return {}

        tag_ids = list(tag_map.values())
        id_to_name = {v: k for k, v in tag_map.items()}
        result = {}

        if aggregation == 'last':
            cur.execute("""
                SELECT DISTINCT ON (h.tag_id) h.tag_id, h.value
                FROM tag_history h
                WHERE h.tag_id = ANY(%s)
                  AND h."timestamp" >= %s::timestamp
                  AND h."timestamp" <= %s::timestamp
                ORDER BY h.tag_id, h."timestamp" DESC
            """, (tag_ids, from_dt, to_dt))
            for row in cur.fetchall():
                name = id_to_name.get(row['tag_id'])
                if name:
                    result[name] = row['value']
        elif aggregation == 'first':
            cur.execute("""
                SELECT DISTINCT ON (h.tag_id) h.tag_id, h.value
                FROM tag_history h
                WHERE h.tag_id = ANY(%s)
                  AND h."timestamp" >= %s::timestamp
                  AND h."timestamp" <= %s::timestamp
                ORDER BY h.tag_id, h."timestamp" ASC
            """, (tag_ids, from_dt, to_dt))
            for row in cur.fetchall():
                name = id_to_name.get(row['tag_id'])
                if name:
                    result[name] = row['value']
        elif aggregation == 'delta':
            cur.execute("""
                SELECT DISTINCT ON (h.tag_id) h.tag_id, h.value
                FROM tag_history h
                WHERE h.tag_id = ANY(%s)
                  AND h."timestamp" >= %s::timestamp
                  AND h."timestamp" <= %s::timestamp
                ORDER BY h.tag_id, h."timestamp" ASC
            """, (tag_ids, from_dt, to_dt))
            first_vals = {row['tag_id']: row['value'] for row in cur.fetchall()}

            cur.execute("""
                SELECT DISTINCT ON (h.tag_id) h.tag_id, h.value
                FROM tag_history h
                WHERE h.tag_id = ANY(%s)
                  AND h."timestamp" >= %s::timestamp
                  AND h."timestamp" <= %s::timestamp
                ORDER BY h.tag_id, h."timestamp" DESC
            """, (tag_ids, from_dt, to_dt))
            for row in cur.fetchall():
                name = id_to_name.get(row['tag_id'])
                first = first_vals.get(row['tag_id'])
                if name and first is not None and row['value'] is not None:
                    result[name] = float(row['value']) - float(first)
        else:
            agg_fn = {'avg': 'AVG', 'min': 'MIN', 'max': 'MAX', 'sum': 'SUM', 'count': 'COUNT'}
            fn = agg_fn.get(aggregation, 'AVG')
            cur.execute(f"""
                SELECT h.tag_id, {fn}(h.value) AS agg_value
                FROM tag_history h
                WHERE h.tag_id = ANY(%s)
                  AND h."timestamp" >= %s::timestamp
                  AND h."timestamp" <= %s::timestamp
                GROUP BY h.tag_id
            """, (tag_ids, from_dt, to_dt))
            for row in cur.fetchall():
                name = id_to_name.get(row['tag_id'])
                if name:
                    result[name] = row['agg_value']

        # ── Fallback to tag_history_archive for tags missing from result ──
        missing_ids = [tid for tid, nm in id_to_name.items() if nm not in result]
        if missing_ids:
            if aggregation == 'last':
                cur.execute("""
                    SELECT DISTINCT ON (a.tag_id) a.tag_id, a.value
                    FROM tag_history_archive a
                    WHERE a.tag_id = ANY(%s)
                      AND a.archive_hour >= %s::timestamp
                      AND a.archive_hour <= %s::timestamp
                    ORDER BY a.tag_id, a.archive_hour DESC
                """, (missing_ids, from_dt, to_dt))
                for row in cur.fetchall():
                    name = id_to_name.get(row['tag_id'])
                    if name and row['value'] is not None:
                        result[name] = row['value']
            elif aggregation == 'first':
                cur.execute("""
                    SELECT DISTINCT ON (a.tag_id) a.tag_id, a.value
                    FROM tag_history_archive a
                    WHERE a.tag_id = ANY(%s)
                      AND a.archive_hour >= %s::timestamp
                      AND a.archive_hour <= %s::timestamp
                    ORDER BY a.tag_id, a.archive_hour ASC
                """, (missing_ids, from_dt, to_dt))
                for row in cur.fetchall():
                    name = id_to_name.get(row['tag_id'])
                    if name and row['value'] is not None:
                        result[name] = row['value']
            elif aggregation == 'delta':
                cur.execute("""
                    SELECT DISTINCT ON (a.tag_id) a.tag_id, a.value
                    FROM tag_history_archive a
                    WHERE a.tag_id = ANY(%s)
                      AND a.archive_hour >= %s::timestamp
                      AND a.archive_hour <= %s::timestamp
                    ORDER BY a.tag_id, a.archive_hour ASC
                """, (missing_ids, from_dt, to_dt))
                arch_first = {row['tag_id']: row['value'] for row in cur.fetchall()}
                cur.execute("""
                    SELECT DISTINCT ON (a.tag_id) a.tag_id, a.value
                    FROM tag_history_archive a
                    WHERE a.tag_id = ANY(%s)
                      AND a.archive_hour >= %s::timestamp
                      AND a.archive_hour <= %s::timestamp
                    ORDER BY a.tag_id, a.archive_hour DESC
                """, (missing_ids, from_dt, to_dt))
                for row in cur.fetchall():
                    name = id_to_name.get(row['tag_id'])
                    first = arch_first.get(row['tag_id'])
                    if name and first is not None and row['value'] is not None:
                        result[name] = float(row['value']) - float(first)
            else:
                agg_fn_map = {'avg': 'AVG', 'min': 'MIN', 'max': 'MAX', 'sum': 'SUM', 'count': 'COUNT'}
                fn = agg_fn_map.get(aggregation, 'AVG')
                cur.execute(f"""
                    SELECT a.tag_id, {fn}(a.value) AS agg_value
                    FROM tag_history_archive a
                    WHERE a.tag_id = ANY(%s)
                      AND a.archive_hour >= %s::timestamp
                      AND a.archive_hour <= %s::timestamp
                    GROUP BY a.tag_id
                """, (missing_ids, from_dt, to_dt))
                for row in cur.fetchall():
                    name = id_to_name.get(row['tag_id'])
                    if name and row['agg_value'] is not None:
                        result[name] = row['agg_value']

    return result


def _fetch_time_series_for_distribution(tag_names, from_dt, to_dt, max_points=500):
    """Load per-tag time arrays for dashboard chart images (mirrors historian /time-series).

    Returns:
        dict[str, list[dict]] — ``{ tagName: [{"t": epoch_ms, "v": float}, ...], ... }``
    """
    if not tag_names:
        return {}
    max_points = max(10, min(int(max_points or 500), 5000))

    get_conn = _get_db_connection()
    with closing(get_conn()) as conn:
        actual_conn = conn._conn if hasattr(conn, '_conn') else conn
        cur = actual_conn.cursor(cursor_factory=RealDictCursor)

        cur.execute(
            "SELECT id, tag_name FROM tags WHERE tag_name = ANY(%s) AND is_active = true",
            (list(tag_names),),
        )
        tag_map = {row['tag_name']: row['id'] for row in cur.fetchall()}
        if not tag_map:
            return {}

        tag_ids = list(tag_map.values())
        id_to_name = {v: k for k, v in tag_map.items()}
        result = {}

        cur.execute(
            """
            SELECT COUNT(*) AS cnt
            FROM tag_history h
            WHERE h.tag_id = ANY(%s)
              AND h."timestamp" >= %s::timestamp
              AND h."timestamp" <= %s::timestamp
            """,
            (tag_ids, from_dt, to_dt),
        )
        total_count = cur.fetchone()['cnt']

        if total_count > 0:
            if total_count <= max_points * len(tag_ids):
                cur.execute(
                    """
                    SELECT h.tag_id, h.value,
                           EXTRACT(EPOCH FROM h."timestamp") * 1000 AS t_ms
                    FROM tag_history h
                    WHERE h.tag_id = ANY(%s)
                      AND h."timestamp" >= %s::timestamp
                      AND h."timestamp" <= %s::timestamp
                    ORDER BY h."timestamp"
                    """,
                    (tag_ids, from_dt, to_dt),
                )
            else:
                cur.execute(
                    """
                    WITH bounds AS (
                        SELECT EXTRACT(EPOCH FROM (%s::timestamp - %s::timestamp)) AS range_secs
                    )
                    SELECT h.tag_id,
                           AVG(h.value) AS value,
                           AVG(EXTRACT(EPOCH FROM h."timestamp")) * 1000 AS t_ms
                    FROM tag_history h, bounds
                    WHERE h.tag_id = ANY(%s)
                      AND h."timestamp" >= %s::timestamp
                      AND h."timestamp" <= %s::timestamp
                    GROUP BY h.tag_id,
                             FLOOR(EXTRACT(EPOCH FROM h."timestamp") / GREATEST(1, bounds.range_secs / %s))
                    ORDER BY t_ms
                    """,
                    (to_dt, from_dt, tag_ids, from_dt, to_dt, max_points),
                )

            for row in cur.fetchall():
                name = id_to_name.get(row['tag_id'])
                if name and row['value'] is not None and row['t_ms'] is not None:
                    result.setdefault(name, []).append({'t': round(row['t_ms']), 'v': float(row['value'])})

        if not result:
            cur.execute(
                """
                SELECT a.tag_id, a.value,
                       EXTRACT(EPOCH FROM a.archive_hour) * 1000 AS t_ms
                FROM tag_history_archive a
                WHERE a.tag_id = ANY(%s)
                  AND a.archive_hour >= %s::timestamp
                  AND a.archive_hour <= %s::timestamp
                ORDER BY a.archive_hour
                """,
                (tag_ids, from_dt, to_dt),
            )
            for row in cur.fetchall():
                name = id_to_name.get(row['tag_id'])
                if name and row['value'] is not None and row['t_ms'] is not None:
                    result.setdefault(name, []).append({'t': round(row['t_ms']), 'v': float(row['value'])})

    return result


_CHART_COLOR_DEFAULTS = ['#2563eb', '#7c3aed', '#0891b2', '#059669', '#d97706', '#dc2626', '#ec4899', '#8b5cf6']


def _collect_dashboard_chart_tag_names(layout_config):
    """Tag names referenced by chart / barchart / piechart series (for time-series fetch)."""
    names = set()
    for _heading, flat_widgets in _dashboard_sections_for_distribution(layout_config):
        for w in flat_widgets:
            wt = w.get('type', '')
            if wt not in ('chart', 'barchart', 'piechart'):
                continue
            cfg = w.get('config') or {}
            for s in _chart_series_list_for_export(cfg):
                s_ds = s.get('dataSource', {}) or {}
                tn = s_ds.get('tagName', '') or s.get('tagName', '')
                if tn:
                    names.add(tn)
    return names


def _dashboard_chart_png_data_uri(w_type, config, label, ts_by_tag, tag_data, from_dt, to_dt):
    """Render chart as PNG data URI for embedding in PDF HTML. Returns None to use text fallback."""
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
    except ImportError:
        logger.warning('matplotlib not installed; dashboard charts fall back to text in PDF')
        return None

    series_list = _chart_series_list_for_export(config or {})
    if not series_list:
        return None

    cfg_colors = (config or {}).get('colors') or []
    title = label or (config or {}).get('title') or 'Chart'

    try:
        from io import BytesIO

        fig, ax = plt.subplots(figsize=(6.8, 2.9), dpi=110)
        fig.patch.set_facecolor('#ffffff')
        ax.set_facecolor('#fafafa')

        if w_type == 'piechart':
            sizes = []
            labels = []
            for i, s in enumerate(series_list):
                s_ds = s.get('dataSource', {}) or {}
                tag = s_ds.get('tagName', '') or s.get('tagName', '')
                if not tag:
                    continue
                raw = tag_data.get(tag)
                if raw is None:
                    continue
                try:
                    v = abs(float(raw))
                except (TypeError, ValueError):
                    continue
                if v <= 0:
                    continue
                sizes.append(v)
                labels.append(s.get('label', tag) or tag)
            if not sizes:
                plt.close(fig)
                return None
            col_cycle = [cfg_colors[i % len(cfg_colors)] if cfg_colors else _CHART_COLOR_DEFAULTS[i % len(_CHART_COLOR_DEFAULTS)]
                           for i in range(len(sizes))]
            ax.pie(sizes, labels=labels, colors=col_cycle, autopct='%1.1f%%', startangle=90,
                   textprops={'fontsize': 8})
            ax.set_title(title, fontsize=10, fontweight='bold', color='#0f172a')
        elif w_type == 'barchart' or (w_type == 'chart' and (config or {}).get('chartType') == 'bar'):
            xs = []
            heights = []
            for i, s in enumerate(series_list):
                s_ds = s.get('dataSource', {}) or {}
                tag = s_ds.get('tagName', '') or s.get('tagName', '')
                if not tag:
                    continue
                raw = tag_data.get(tag)
                if raw is None:
                    continue
                try:
                    heights.append(float(raw))
                except (TypeError, ValueError):
                    continue
                xs.append(s.get('label', tag) or tag)
            if not xs:
                plt.close(fig)
                return None
            colors = [cfg_colors[i % len(cfg_colors)] if cfg_colors else _CHART_COLOR_DEFAULTS[i % len(_CHART_COLOR_DEFAULTS)]
                      for i in range(len(xs))]
            ax.bar(range(len(xs)), heights, color=colors, edgecolor='#e2e8f0', linewidth=0.5)
            ax.set_xticks(range(len(xs)))
            ax.set_xticklabels(xs, rotation=25, ha='right', fontsize=8)
            ax.set_ylabel('Value', fontsize=8, color='#64748b')
            ax.grid(True, axis='y', linestyle='--', alpha=0.35)
            ax.set_title(title, fontsize=10, fontweight='bold', color='#0f172a')
        else:
            # Line chart (default)
            plotted = False
            for i, s in enumerate(series_list):
                s_ds = s.get('dataSource', {}) or {}
                tag = s_ds.get('tagName', '') or s.get('tagName', '')
                if not tag:
                    continue
                pts = ts_by_tag.get(tag) or []
                if not pts:
                    continue
                if len(pts) == 1:
                    p0 = pts[0]
                    pts = [
                        p0,
                        {'t': int(p0['t']) + 3600000, 'v': float(p0['v'])},
                    ]
                xdt = [mdates.date2num(datetime.fromtimestamp(float(p['t']) / 1000.0)) for p in pts]
                vs = [float(p['v']) for p in pts]
                color = (s.get('color') or (cfg_colors[i % len(cfg_colors)] if cfg_colors
                        else _CHART_COLOR_DEFAULTS[i % len(_CHART_COLOR_DEFAULTS)]))
                ax.plot(xdt, vs, linewidth=1.6, label=s.get('label', tag) or tag, color=color)
                plotted = True
            if not plotted:
                plt.close(fig)
                return None
            ax.xaxis.set_major_formatter(mdates.DateFormatter('%d/%m %H:%M'))
            ax.tick_params(axis='x', labelsize=7, rotation=18)
            ax.tick_params(axis='y', labelsize=8)
            ax.grid(True, linestyle='--', alpha=0.35)
            ax.set_title(title, fontsize=10, fontweight='bold', color='#0f172a')
            if len(series_list) > 1:
                ax.legend(loc='upper right', fontsize=7, framealpha=0.92)

        fig.tight_layout()
        buf = BytesIO()
        fig.savefig(buf, format='png', bbox_inches='tight', facecolor=fig.get_facecolor())
        plt.close(fig)
        buf.seek(0)
        b64 = base64.standard_b64encode(buf.read()).decode('ascii')
        return f'data:image/png;base64,{b64}'
    except Exception as e:
        logger.warning('Dashboard chart render failed (%s): %s', title, e)
        try:
            plt.close('all')
        except Exception:
            pass
        return None


# ── Safe string helper ───────────────────────────────────────────────────────

def _esc(val):
    """HTML-escape a value for safe embedding in generated reports."""
    return html_escape(str(val))


# ── Formula evaluator (mirrors frontend formulaEngine.js) ────────────────────

def _safe_aggregate(fn_name, args_str):
    """Safely compute aggregate functions without eval()."""
    try:
        nums = [float(x.strip()) for x in args_str.split(',') if x.strip()]
    except (ValueError, TypeError):
        return '0'
    if not nums:
        return '0'
    if fn_name == 'sum':
        return str(sum(nums))
    elif fn_name == 'avg':
        return str(sum(nums) / len(nums))
    elif fn_name == 'min':
        return str(min(nums))
    elif fn_name == 'max':
        return str(max(nums))
    return '0'


# Thread-safe asteval interpreter — create per call to avoid shared state
def _get_formula_interp():
    from asteval import Interpreter
    return Interpreter()


def _count_bare_formula_tag_occurrences(formula):
    """Count bare {Tag} refs (no agg:: prefix) per base name — used for {T}-{T} totalizer delta."""
    counts = {}
    for base, explicit in _parse_formula_tag_tokens(formula):
        if explicit:
            continue
        counts[base] = counts.get(base, 0) + 1
    return counts


def _evaluate_formula(formula, tag_data, aggregation=None):
    """Evaluate a formula string like '{Tag1} + {Tag2} * 100'. Returns float or None.

    Mirrors frontend formulaEngine.js: supports {agg::Tag}; when the same bare {Tag}
    appears exactly twice, first occurrence is last-in-range and the second is first
    (so '{T}-{T}' equals delta for paginated historian merges).
    """
    if not formula or not formula.strip():
        return None
    try:
        bare_dup = _count_bare_formula_tag_occurrences(formula)
        bare_occ = {k: 0 for k in bare_dup}

        def _resolve_tag(m):
            inner = m.group(1)
            if inner.startswith('col:'):
                return '0'
            mm = _FORMULA_AGG_PREFIX_RE.match(inner)
            if mm:
                agg, base = mm.group(1).lower(), mm.group(2)
                key = base if agg == 'last' else f'{agg}::{base}'
                val = tag_data.get(key)
                if val is None and agg == 'last':
                    val = tag_data.get(base)
                return str(float(val)) if val is not None else '0'
            name = inner
            use_dup = bare_dup.get(name) == 2
            if use_dup:
                k = bare_occ[name]
                bare_occ[name] = k + 1
                if k == 0:
                    val = tag_data.get(name)
                    if val is None:
                        val = tag_data.get(f'last::{name}')
                else:
                    val = tag_data.get(f'first::{name}')
                    if val is None:
                        val = tag_data.get(name)
                return str(float(val)) if val is not None else '0'
            if aggregation and aggregation != 'last':
                key = f'{aggregation}::{name}'
                if key in tag_data:
                    return str(float(tag_data[key]))
            val = tag_data.get(name)
            return str(float(val)) if val is not None else '0'

        expr = re.sub(r'\{([^}]+)\}', _resolve_tag, formula)

        for fn_pattern, fn_safe in [('SUM', 'sum'), ('AVG', 'avg'), ('MIN', 'min'), ('MAX', 'max')]:
            expr = re.sub(
                rf'\b{fn_pattern}\s*\(([^)]*)\)',
                lambda m, f=fn_safe: _safe_aggregate(f, m.group(1)),
                expr, flags=re.IGNORECASE
            )
        expr = re.sub(r'\bABS\s*\(([^)]*)\)', lambda m: str(abs(float(m.group(1)))), expr, flags=re.IGNORECASE)
        expr = re.sub(r'\bROUND\s*\(([^,]*),([^)]*)\)', lambda m: str(round(float(m.group(1)), int(m.group(2)))), expr, flags=re.IGNORECASE)
        expr = re.sub(r'\bIF\s*\(([^,]*),([^,]*),([^)]*)\)', lambda m: m.group(2).strip() if float(m.group(1)) else m.group(3).strip(), expr, flags=re.IGNORECASE)
        expr = re.sub(r'\bCLAMP\s*\(([^,]*),([^,]*),([^)]*)\)', lambda m: str(max(float(m.group(2)), min(float(m.group(3)), float(m.group(1))))), expr, flags=re.IGNORECASE)
        expr = re.sub(r'\bDIFF\s*\(([^)]*)\)', lambda m: m.group(1), expr, flags=re.IGNORECASE)
        expr = re.sub(r'\bRATE\s*\(([^)]*)\)', lambda m: m.group(1), expr, flags=re.IGNORECASE)

        sanitized = re.sub(r'[^0-9+\-*/%().eE\s]', '', expr)
        if not sanitized.strip():
            return None
        interp = _get_formula_interp()
        result = interp(sanitized)
        return float(result) if isinstance(result, (int, float)) and not (isinstance(result, float) and (result != result)) else None
    except Exception:
        return None


def _resolve_widget_datasource_value(ds, tag_data):
    """Scalar from a widget dataSource (tag / formula / group) for static export."""
    if not ds:
        return None
    ds_type = ds.get('type', 'tag') or 'tag'
    if ds_type == 'formula' and ds.get('formula'):
        return _evaluate_formula(ds['formula'], tag_data)
    if ds_type == 'group' and ds.get('groupTags'):
        vals = []
        for t in ds['groupTags']:
            v = tag_data.get(t)
            if v is not None:
                try:
                    vals.append(float(v))
                except (TypeError, ValueError):
                    pass
        if not vals:
            return None
        agg = ds.get('aggregation', 'avg')
        if agg == 'sum':
            return sum(vals)
        if agg == 'min':
            return min(vals)
        if agg == 'max':
            return max(vals)
        if agg == 'count':
            return len(vals)
        return sum(vals) / len(vals)
    tn = ds.get('tagName', '')
    if tn:
        return tag_data.get(tn)
    return None


# ── Resolve a paginated cell value ───────────────────────────────────────────

def _resolve_cell(cell, tag_data):
    """Resolve a single paginated cell to a display string."""
    if not cell:
        return ''
    src = cell.get('sourceType', 'static')
    raw_dec = cell.get('decimals')
    decimals = int(raw_dec) if raw_dec is not None and raw_dec != '' else 0
    decimals = max(0, min(10, decimals))
    is_checkbox = cell.get('unit') == '__checkbox__'
    unit = cell.get('unit', '')
    if unit == '__checkbox__':
        unit = ''
    elif unit == '__custom__':
        unit = cell.get('customUnit', '')

    def _fmt_check(val):
        """Format checkbox cells — same wording as paginated Report Builder (Yes / No)."""
        if val is None:
            return '—'
        try:
            n = float(val)
            return 'Yes' if n == 1 else 'No'
        except (TypeError, ValueError):
            s = str(val).strip().lower()
            if s in ('1', 'true', 'yes', 'y', 'on'):
                return 'Yes'
            if s in ('0', 'false', 'no', 'n', 'off'):
                return 'No'
            return 'Yes' if s else 'No'

    def _fmt(val):
        if val is None:
            return '—'
        if is_checkbox:
            return _fmt_check(val)
        try:
            n = float(val)
            formatted = f"{n:,.{decimals}f}"
            return f"{formatted}\u00a0{unit}" if unit else formatted
        except (TypeError, ValueError):
            return str(val)

    if src == 'static':
        return cell.get('value', '')

    if src == 'tag':
        tag_name = cell.get('tagName', '')
        agg = cell.get('aggregation', 'last')
        key = f'{agg}::{tag_name}' if agg and agg != 'last' else tag_name
        raw = tag_data.get(key)
        return _fmt(raw)

    if src == 'formula':
        result = _evaluate_formula(cell.get('formula', ''), tag_data, aggregation=cell.get('aggregation'))
        return _fmt(result)

    if src == 'group':
        vals = [float(tag_data.get(t, 0)) for t in (cell.get('groupTags') or []) if tag_data.get(t) is not None]
        if not vals:
            return '—'
        agg = cell.get('aggregation', 'avg')
        if agg == 'sum':
            n = sum(vals)
        elif agg == 'min':
            n = min(vals)
        elif agg == 'max':
            n = max(vals)
        elif agg == 'count':
            n = len(vals)
        else:
            n = sum(vals) / len(vals)
        return _fmt(n)

    return '—'


def _resolve_cell_raw(cell, tag_data):
    """Resolve a single cell to a raw (value, unit, decimals) tuple for Excel output.
    Returns: (value: float|str|None, unit: str, decimals: int)
    - value is a raw float for numeric cells (no formatting), str for static, None for missing
    - Checkbox cells return 'Yes' or 'No' (or '—' if unset)
    """
    if not cell:
        return (None, '', 0)
    src = cell.get('sourceType', 'static')
    raw_dec = cell.get('decimals')
    decimals = int(raw_dec) if raw_dec is not None and raw_dec != '' else 0
    decimals = max(0, min(10, decimals))
    is_checkbox = cell.get('unit') == '__checkbox__'
    unit = cell.get('unit', '')
    if unit == '__checkbox__':
        unit = ''
    elif unit == '__custom__':
        unit = cell.get('customUnit', '')

    def _to_check(val):
        """Convert value to Yes/No for checkbox columns (Excel export)."""
        if val is None:
            return '—'
        try:
            n = float(val)
            return 'Yes' if n == 1 else 'No'
        except (TypeError, ValueError):
            s = str(val).strip().lower()
            if s in ('1', 'true', 'yes', 'y', 'on'):
                return 'Yes'
            if s in ('0', 'false', 'no', 'n', 'off'):
                return 'No'
            return 'Yes' if s else 'No'

    def _to_num(val):
        if val is None:
            return None
        try:
            return float(val)
        except (TypeError, ValueError):
            return str(val)

    if src == 'static':
        v = cell.get('value', '')
        try:
            return (float(v), unit, decimals)
        except (TypeError, ValueError):
            return (v, unit, decimals)

    if src == 'tag':
        tag_name = cell.get('tagName', '')
        agg = cell.get('aggregation', 'last')
        key = f'{agg}::{tag_name}' if agg and agg != 'last' else tag_name
        raw = tag_data.get(key)
        if is_checkbox:
            return (_to_check(raw), '', 0)
        return (_to_num(raw), unit, decimals)

    if src == 'formula':
        result = _evaluate_formula(cell.get('formula', ''), tag_data, aggregation=cell.get('aggregation'))
        if is_checkbox:
            return (_to_check(result), '', 0)
        return (_to_num(result), unit, decimals)

    if src == 'group':
        vals = [float(tag_data.get(t, 0)) for t in (cell.get('groupTags') or []) if tag_data.get(t) is not None]
        if not vals:
            return (None, unit, decimals)
        agg = cell.get('aggregation', 'avg')
        if agg == 'sum':
            n = sum(vals)
        elif agg == 'min':
            n = min(vals)
        elif agg == 'max':
            n = max(vals)
        elif agg == 'count':
            n = len(vals)
        else:
            n = sum(vals) / len(vals)
        if is_checkbox:
            return (_to_check(n), '', 0)
        return (n, unit, decimals)

    return (None, '', 1)


# ── Row visibility (mirrors frontend isRowHidden) ────────────────────────────

def _is_row_hidden(row, tag_data):
    """Return True if the row should be hidden (bin inactive)."""
    if not row.get('hideWhenInactive'):
        return False
    ref_col = row.get('hideReferenceCol', 0) or 0
    cells = row.get('cells', [])
    if ref_col >= len(cells):
        return False
    cell = cells[ref_col]
    if not cell:
        return False
    resolved = _resolve_cell(cell, tag_data)
    if resolved in ('—', ''):
        return True
    cleaned = re.sub(r'[^0-9.\-]', '', str(resolved))
    if cleaned:
        try:
            return float(cleaned) == 0
        except (TypeError, ValueError):
            pass
    return False


# ── Logo helpers ─────────────────────────────────────────────────────────────

def _file_to_base64_data_uri(filepath):
    """Read an image file and return a data:image URI for HTML embedding."""
    if not os.path.isfile(filepath):
        return ''
    ext = os.path.splitext(filepath)[1].lower()
    mime = {'png': 'image/png', 'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
            'svg': 'image/svg+xml', 'gif': 'image/gif'}.get(ext.lstrip('.'), 'image/png')
    with open(filepath, 'rb') as f:
        encoded = base64.b64encode(f.read()).decode('ascii')
    return f'data:{mime};base64,{encoded}'


def _get_logo_data_uris():
    """Return (hercules_uri, asm_uri, client_logo_uri) for embedding in HTML reports."""
    backend_dir = os.path.abspath(os.path.dirname(__file__))
    project_root = os.path.dirname(backend_dir)

    # Search multiple possible logo directories (Vite dist, static, source assets)
    # PyInstaller bundles to _internal/frontend/dist/assets/ (lowercase)
    search_dirs = [
        os.path.join(backend_dir, 'frontend', 'dist', 'assets'),    # PyInstaller: _internal/frontend/dist/assets/
        os.path.join(backend_dir, 'static', 'assets'),              # backend/static/assets/
        os.path.join(project_root, 'Frontend', 'dist', 'assets'),   # Frontend/dist/assets/
        os.path.join(project_root, 'Frontend', 'src', 'Assets'),    # Frontend/src/Assets/
        os.path.join(project_root, 'frontend', 'dist', 'assets'),   # frontend/dist/assets/ (lowercase)
    ]

    hercules_uri = ''
    asm_uri = ''
    for search_dir in search_dirs:
        if not os.path.isdir(search_dir):
            continue
        for fname in os.listdir(search_dir):
            if fname.startswith('Hercules_New') and fname.endswith('.png') and not hercules_uri:
                hercules_uri = _file_to_base64_data_uri(os.path.join(search_dir, fname))
            elif fname.startswith('Asm_Logo') and fname.endswith('.png') and not asm_uri:
                asm_uri = _file_to_base64_data_uri(os.path.join(search_dir, fname))
        if hercules_uri and asm_uri:
            break  # Found both, stop searching

    # Client logo from DB (stored as base64 data URI)
    client_logo_uri = ''
    try:
        get_conn = _get_db_connection()
        with closing(get_conn()) as conn:
            actual_conn = conn._conn if hasattr(conn, '_conn') else conn
            cur = actual_conn.cursor(cursor_factory=RealDictCursor)
            cur.execute("SELECT value FROM system_settings WHERE key = 'client_logo'")
            row = cur.fetchone()
            if row and row.get('value'):
                client_logo_uri = row['value']
    except Exception:
        pass

    return hercules_uri, asm_uri, client_logo_uri


# ── Shared CSS (matches frontend PaginatedReportPreview / ReportViewer) ──────

_SHARED_CSS = """
@page { size: A4; margin: 6mm 10mm; }
body {
  font-family: Inter, system-ui, -apple-system, sans-serif;
  font-size: 12px;
  color: #1a1a2e;
  line-height: 1.35;
  margin: 0;
  padding: 0;
  -webkit-print-color-adjust: exact;
  print-color-adjust: exact;
}

/* ── Logo header bar ── */
.logo-header-table { margin-bottom: 2px; border-bottom: 1.5px solid #e2e8f0; padding-bottom: 2px; }
.logo-header-table img { vertical-align: middle; }

/* ── Report header ── */
h1.report-title {
  font-size: 18px;
  font-weight: 700;
  letter-spacing: -0.02em;
  color: #0f172a;
  margin: 0 0 1px 0;
}
p.subtitle { font-size: 12px; color: #64748b; margin: 0 0 1px 0; }
p.period   { font-size: 12px; color: #64748b; font-weight: 500; margin: 0 0 2px 0; }
.header-rule {
  margin-top: 2px;
  height: 2px;
  background-color: #1a5276;
}

/* ── Section label ── */
.section-label {
  font-size: 12px;
  font-weight: 700;
  color: #0f172a;
  margin: 6px 0 2px 0;
}
.kpi-section-label {
  font-size: 10px;
  font-weight: 700;
  text-transform: uppercase;
  letter-spacing: 0.05em;
  color: #94a3b8;
  margin: 6px 0 2px 0;
}

/* ── KPI row ── */
.kpi-row-table { width: 100%; margin-bottom: 4px; }
.kpi-row-table td { text-align: right; padding: 1px 8px; }
.kpi-label { font-size: 10px; font-weight: 500; color: #64748b; }
.kpi-value { font-size: 12px; font-weight: 700; color: #0f172a; }

/* ── Data tables (paginated PDF) ──
   xhtml2pdf aligns thead/tbody poorly if only <th> has widths; use <colgroup>
   plus matching width on <td>. Avoid flex in table rows. */
table.data-table {
  width: 100%;
  border-collapse: collapse;
  table-layout: fixed;
  margin: 2px 0 6px 0;
  font-size: 11px;
}
table.data-table th {
  padding: 5px 8px;
  font-weight: 700;
  font-size: 10px;
  border: 1px solid #94a3b8;
  background-color: #1e293b;
  color: #ffffff;
  text-transform: uppercase;
  letter-spacing: 0.03em;
  overflow: hidden;
  word-wrap: break-word;
}
table.data-table td {
  padding: 4px 8px;
  font-size: 11px;
  border: 1px solid #d1d5db;
  color: #1e293b;
  overflow: hidden;
  word-wrap: break-word;
}
table.data-table .alt-row { background-color: #f1f5f9; }
table.data-table .summary-row { font-weight: 700; background-color: #e0f2fe; }
table.data-table .summary-row td { border-top: 2px solid #94a3b8; font-size: 11px; }

/* ── Text blocks ── */
.text-block { margin-bottom: 4px; }

/* ── Signature block ── */
.sig-block { margin-top: 16px; margin-bottom: 6px; }
.sig-block table { width: 100%; border-collapse: collapse; }
.sig-block td { padding: 0 16px; vertical-align: top; }
.sig-label { font-size: 11px; font-weight: 500; color: #64748b; margin-bottom: 20px; }
.sig-line  { border-bottom: 1px solid #cbd5e1; padding-bottom: 4px; font-size: 12px; color: #334155; min-height: 18px; }
.sig-date  { font-size: 10px; color: #94a3b8; margin-top: 4px; }

/* ── Page footer ── */
.page-footer {
  position: fixed;
  bottom: 0;
  left: 0;
  right: 0;
  display: flex;
  justify-content: space-between;
  padding: 2px 10mm;
  font-size: 9px;
  color: #94a3b8;
}

/* ── Dashboard grid (align with Report Viewer / PDF export look) ── */
.dashboard-section { margin-bottom: 8px; }
.dashboard-tab-block {
  margin-top: 4px;
  margin-bottom: 10px;
}
.dashboard-tab-block.tab-after-first {
  page-break-before: always;
}
/* Tab title bar — table layout for xhtml2pdf (avoids flex issues) */
table.report-tab-banner {
  width: 100%;
  border-collapse: collapse;
  margin: 10px 0 0 0;
}
table.report-tab-banner td {
  background-color: #1a3a5c;
  color: #ffffff;
  font-size: 15px;
  font-weight: 700;
  letter-spacing: 0.02em;
  padding: 8px 14px;
  vertical-align: middle;
}
.report-tab-accent {
  height: 3px;
  background-color: #0f3460;
  margin: 0 0 10px 0;
}
.report-section-title {
  font-size: 14px;
  font-weight: 700;
  color: #1a5276;
  margin: 14px 0 6px 0;
  padding: 4px 0;
  border-bottom: 2px solid #1a5276;
}
.dashboard-card {
  background: #ffffff;
  border: 1px solid #d1d5db;
  border-radius: 4px;
  padding: 8px 12px;
  margin-bottom: 6px;
}
.widget-label { font-size: 9px; font-weight: 700; color: #64748b; text-transform: uppercase; letter-spacing: 0.04em; margin-bottom: 2px; }
.widget-value { font-size: 16px; font-weight: 700; color: #0f172a; }
.widget-unit  { font-size: 10px; font-weight: 500; color: #94a3b8; margin-left: 3px; }
.widget-silo  { font-size: 12px; color: #334155; }
.widget-chart-note { font-size: 9px; color: #94a3b8; font-style: italic; padding: 2px 0; }
.dashboard-card img {
  max-width: 100%;
  height: auto;
  display: block;
  margin-top: 4px;
}

/* ── Dashboard grid table ── */
table.grid-row { width: 100%; border-collapse: collapse; margin-bottom: 2px; }
table.grid-row td { vertical-align: top; padding: 0 1px; }

/* ── Generated footer ── */
.gen-footer {
  margin-top: 12px;
  font-size: 9px;
  color: #94a3b8;
  border-top: 1px solid #e5e7eb;
  padding-top: 4px;
}
table.gen-footer-table {
  width: 100%;
  border-collapse: collapse;
  margin-top: 12px;
  font-size: 9px;
  color: #94a3b8;
  border-top: 1px solid #e5e7eb;
}
table.gen-footer-table td {
  padding-top: 4px;
  vertical-align: top;
}
"""


# ── Logo header HTML builder ─────────────────────────────────────────────────

def _build_logo_header_html(hercules_uri, asm_uri, client_logo_uri):
    """Build the logo header bar using table layout (xhtml2pdf doesn't support flexbox)."""
    herc_td = f'<img src="{hercules_uri}" alt="Hercules" style="height:44px;width:auto" />' if hercules_uri else '&nbsp;'
    right_imgs = ''
    if client_logo_uri:
        right_imgs += f'<img src="{client_logo_uri}" alt="Client" style="height:40px;width:auto;max-width:140px" /> '
    if asm_uri:
        right_imgs += f'<img src="{asm_uri}" alt="ASM" style="height:40px;width:auto" />'
    right_td = right_imgs or '&nbsp;'

    return f"""<table width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:4px;border-bottom:1.5px solid #e2e8f0;padding-bottom:4px">
  <tr>
    <td style="text-align:left;vertical-align:middle">{herc_td}</td>
    <td style="text-align:right;vertical-align:middle">{right_td}</td>
  </tr>
</table>"""


# ── HTML report generation ───────────────────────────────────────────────────

GRID_COLS = 12  # react-grid-layout uses 12 columns


def _render_single_widget_html(widget, tag_data, ts_by_tag=None, from_dt=None, to_dt=None):
    """Render one dashboard widget to an HTML string."""
    w_type = widget.get('type', '')
    config = widget.get('config', {}) or {}
    label = config.get('label') or config.get('title') or widget.get('i', w_type)

    if w_type in ('kpi', 'gauge', 'stat', 'progress', 'sparkline'):
        ds = config.get('dataSource', {}) or {}
        raw = _resolve_widget_datasource_value(ds, tag_data)
        if raw is None:
            val = '—'
        elif isinstance(raw, float):
            val = f"{raw:,.2f}"
        else:
            val = str(raw)
        unit = config.get('unit', '')
        unit_html = f'<span class="widget-unit">{_esc(unit)}</span>' if unit else ''
        return f'<div class="dashboard-card"><div class="widget-label">{_esc(label)}</div><div class="widget-value">{_esc(val)}{unit_html}</div></div>\n'

    if w_type == 'silo':
        ds = config.get('dataSource', {}) or {}
        tag = ds.get('tagName', '')
        s_val = tag_data.get(tag, '—')
        if isinstance(s_val, float):
            s_val = f"{s_val:.1f}"
        cap_tag = config.get('capacityTag', '')
        cap_val = tag_data.get(cap_tag, '—')
        tons_tag = config.get('tonsTag', '')
        tons_val = tag_data.get(tons_tag, '—')
        if isinstance(cap_val, float):
            cap_val = f"{cap_val:,.0f}"
        if isinstance(tons_val, float):
            tons_val = f"{tons_val:,.1f}"
        return (
            f'<div class="dashboard-card"><div class="widget-label">{_esc(label)}</div>'
            f'<div class="widget-silo">Level: <strong>{_esc(s_val)}%</strong> &nbsp;|&nbsp; '
            f'Capacity: <strong>{_esc(cap_val)}</strong> &nbsp;|&nbsp; Tons: <strong>{_esc(tons_val)}</strong></div></div>\n'
        )

    if w_type in ('chart', 'barchart', 'piechart'):
        # Try time-series chart image via the dedicated renderer
        img_uri = _dashboard_chart_png_data_uri(w_type, config, label, ts_by_tag or {}, tag_data, from_dt, to_dt) if from_dt else None
        if img_uri:
            return (
                f'<div class="dashboard-card"><div class="widget-label">{_esc(label)}</div>'
                f'<img src="{img_uri}" alt="" />'
                f'</div>\n'
            )
        # Fallback: render series as bar chart or value table
        series_data = []
        for s in _chart_series_list_for_export(config):
            s_ds = s.get('dataSource', {}) or {}
            s_tag = s_ds.get('tagName', '') or s.get('tagName', '')
            s_raw = tag_data.get(s_tag) if s_tag else None
            s_label = s.get('label', s_tag or 'Series')
            try:
                s_val = float(s_raw) if s_raw is not None else 0
            except (TypeError, ValueError):
                s_val = 0
            series_data.append((s_label, s_val))
        w_grid = widget.get('w', 6)
        chart_uri = _render_chart_image_base64(w_type, series_data, label, width_px=w_grid * 45, height_px=160)
        if chart_uri:
            return f'<div class="dashboard-card"><div class="widget-label">{_esc(label)}</div><img src="{chart_uri}" alt="" /></div>\n'
        # Last resort: text table
        parts = []
        for sl, sv in series_data:
            parts.append(f'<tr><td style="padding:2px 8px;font-size:11px">{_esc(sl)}</td><td style="padding:2px 8px;font-weight:700;text-align:right;font-size:11px">{sv:,.2f}</td></tr>')
        return (
            f'<div class="dashboard-card"><div class="widget-label">{_esc(label)}</div>'
            f'<table style="width:100%;border-collapse:collapse">{"".join(parts)}</table></div>\n'
        )

    if w_type == 'table':
        cols = config.get('tableColumns') or config.get('columns') or []
        header_cells = ''.join(f'<th>{_esc(c.get("label", ""))}</th>' for c in cols)
        data_cells = ''
        for col in cols:
            col_type = col.get('sourceType', 'tag')
            col_agg = col.get('aggregation', 'last')
            if col_type == 'tag':
                tag_key = f'{col_agg}::{col.get("tagName", "")}' if col_agg and col_agg != 'last' else col.get('tagName', '')
                c_val = tag_data.get(tag_key, tag_data.get(col.get('tagName', ''), '—'))
                if isinstance(c_val, float):
                    c_val = f"{c_val:,.2f}"
                data_cells += f'<td>{_esc(c_val)}</td>'
            elif col_type == 'group':
                vals = [str(tag_data.get(t, '—')) for t in col.get('groupTags', [])]
                data_cells += f'<td>{_esc(", ".join(vals))}</td>'
            elif col_type == 'formula':
                result = _evaluate_formula(col.get('formula', ''), tag_data, aggregation=col_agg)
                data_cells += f'<td>{_esc(f"{result:,.2f}" if isinstance(result, float) else (result or "—"))}</td>'
            else:
                data_cells += '<td>—</td>'
        return (
            f'<div class="dashboard-card"><div class="widget-label">{_esc(label)}</div>'
            f'<table class="data-table"><thead><tr>{header_cells}</tr></thead><tbody><tr>{data_cells}</tr></tbody></table></div>\n'
        )

    if w_type == 'datapanel':
        fields = config.get('fields', [])
        field_parts = []
        for f in fields:
            f_label = f.get('label', '')
            src = f.get('sourceType', 'static')
            if src == 'static':
                f_val = f.get('value', '')
            elif src == 'tag' and f.get('tagName'):
                f_val = tag_data.get(f['tagName'], '—')
                if isinstance(f_val, float):
                    dec = f.get('decimals', 2)
                    f_val = f"{f_val:,.{dec}f}"
                f_unit = f.get('unit', '')
                if f_unit:
                    f_val = f"{f_val} {f_unit}"
            elif src == 'formula' and f.get('formula'):
                f_val = _evaluate_formula(f['formula'], tag_data)
                if isinstance(f_val, (int, float)):
                    f_val = f"{f_val:,.2f}"
                elif f_val is None:
                    f_val = '—'
            else:
                f_val = '—'
            if f_label or f_val:
                field_parts.append(
                    f'<tr><td style="padding:2px 8px;color:#64748b;font-size:10px">{_esc(f_label)}</td>'
                    f'<td style="padding:2px 8px;font-weight:600;text-align:right">{_esc(f_val)}</td></tr>'
                )
        if field_parts:
            return (
                f'<div class="dashboard-card"><div class="widget-label">{_esc(label)}</div>'
                f'<table style="width:100%;font-size:11px;border-collapse:collapse">{"".join(field_parts)}</table></div>\n'
            )
        return ''

    if w_type == 'statusbar':
        status_parts = []
        for st in config.get('tags', []):
            if not isinstance(st, dict):
                continue
            st_tag = st.get('tagName', '')
            raw = tag_data.get(st_tag)
            num = float(raw) if raw is not None else None
            is_on = num == 1 if num is not None else False
            status_text = st.get('onLabel', 'ON') if is_on else st.get('offLabel', 'OFF')
            dot = '●' if is_on else '○'
            st_label = st.get('label', st_tag)
            status_parts.append(f'{_esc(st_label)}: {dot} {_esc(status_text)}')
        if status_parts:
            return f'<div class="dashboard-card"><div class="widget-label">{_esc(label)}</div><div style="font-size:11px">{" &nbsp;|&nbsp; ".join(status_parts)}</div></div>\n'
        return ''

    if w_type == 'status':
        ds = config.get('dataSource', {}) or {}
        tag = ds.get('tagName', '')
        raw = tag_data.get(tag)
        num = float(raw) if raw is not None else None
        status_text = '—'
        for zone in config.get('zones', []):
            if num is not None and zone.get('from', 0) <= num <= zone.get('to', 0):
                status_text = zone.get('status', str(num))
                break
        return f'<div class="dashboard-card"><div class="widget-label">{_esc(label)}</div><div class="widget-value" style="font-size:16px">{_esc(status_text)}</div></div>\n'

    if w_type == 'hopper':
        ds = config.get('dataSource', {}) or {}
        tag = ds.get('tagName', '')
        val = tag_data.get(tag, '—')
        if isinstance(val, float):
            val = f"{val:.1f}"
        cap_tag = config.get('capacityTag', '')
        cap_val = tag_data.get(cap_tag, '—') if cap_tag else '—'
        if isinstance(cap_val, float):
            cap_val = f"{cap_val:,.0f}"
        unit = config.get('unit', '%')
        html = f'<div class="dashboard-card"><div class="widget-label">{_esc(label)}</div>'
        html += f'<div class="widget-silo">Level: <strong>{_esc(val)} {_esc(unit)}</strong>'
        if cap_tag:
            html += f' &nbsp;|&nbsp; Capacity: <strong>{_esc(cap_val)}</strong>'
        html += '</div></div>\n'
        return html

    if w_type in ('text', 'header', 'image', 'spacer', 'divider', 'logo'):
        if w_type == 'text':
            text = config.get('text', config.get('content', ''))
            if text:
                return f'<div class="text-block" style="font-size:13px;color:#334155">{_esc(text)}</div>\n'
        elif w_type == 'header':
            text = config.get('text', config.get('title', ''))
            if text:
                return f'<div class="text-block" style="font-size:16px;font-weight:700;color:#0f172a">{_esc(text)}</div>\n'
        return ''

    # Generic fallback
    ds = config.get('dataSource', {}) or {}
    raw = _resolve_widget_datasource_value(ds, tag_data)
    if raw is not None:
        val = f"{raw:,.2f}" if isinstance(raw, float) else str(raw)
        unit = config.get('unit', '')
        unit_html = f'<span class="widget-unit">{_esc(unit)}</span>' if unit else ''
        return f'<div class="dashboard-card"><div class="widget-label">{_esc(label)}</div><div class="widget-value">{_esc(val)}{unit_html}</div></div>\n'
    return ''


def _generate_dashboard_html(report_name, layout_config, tag_data, from_dt, to_dt):
    """Generate HTML report from dashboard widgets with grid layout."""
    period_start = from_dt.strftime('%d/%m/%Y, %H:%M')
    period_end = to_dt.strftime('%d/%m/%Y, %H:%M')
    hercules_uri, asm_uri, client_logo_uri = _get_logo_data_uris()

    chart_tag_names = _collect_dashboard_chart_tag_names(layout_config)
    ts_by_tag = (
        _fetch_time_series_for_distribution(chart_tag_names, from_dt, to_dt)
        if chart_tag_names
        else {}
    )

    cards_html = ""
    for tab_index, (section_heading, flat_widgets) in enumerate(_dashboard_sections_for_distribution(layout_config)):
        tab_cls = "dashboard-tab-block"
        if tab_index > 0:
            tab_cls += " tab-after-first"
        cards_html += f'<div class="{tab_cls}">\n'
        if section_heading:
            cards_html += (
                '<table class="report-tab-banner" width="100%" cellpadding="0" cellspacing="0"><tr>'
                f'<td>{_esc(section_heading)}</td>'
                '</tr></table>\n'
                '<div class="report-tab-accent"></div>\n'
            )

        sorted_widgets = sorted(flat_widgets, key=lambda w: (w.get('y', 0), w.get('x', 0)))

        # Group widgets into rows by y-coordinate
        rows = {}
        for widget in sorted_widgets:
            y = widget.get('y', 0)
            rows.setdefault(y, []).append(widget)

        for y_key in sorted(rows.keys()):
            row_widgets = sorted(rows[y_key], key=lambda w: w.get('x', 0))

            # Check if all widgets span the full width (full-width row)
            total_w = sum(w.get('w', GRID_COLS) for w in row_widgets)
            single_full = len(row_widgets) == 1 and row_widgets[0].get('w', GRID_COLS) >= GRID_COLS

            if single_full or len(row_widgets) == 1:
                # Full-width: render directly
                cards_html += _render_single_widget_html(row_widgets[0], tag_data, ts_by_tag, from_dt, to_dt)
            else:
                # Multi-column row: normalize widths to 100% of the row
                row_total_w = sum(w.get('w', 3) for w in row_widgets) or 1
                cards_html += '<table class="grid-row"><tr>\n'
                for widget in row_widgets:
                    w_span = widget.get('w', 3)
                    width = f'{(w_span / row_total_w) * 100:.1f}%'
                    cell_html = _render_single_widget_html(widget, tag_data, ts_by_tag, from_dt, to_dt)
                    cards_html += f'<td style="width:{width}">{cell_html}</td>\n'
                cards_html += '</tr></table>\n'

        cards_html += '</div>\n'

    logo_html = _build_logo_header_html(hercules_uri, asm_uri, client_logo_uri)

    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><style>{_SHARED_CSS}</style></head>
<body style="padding: 3mm 8mm 5mm 8mm; width: 194mm;">
{logo_html}
<h1 class="report-title" style="text-align:center">{_esc(report_name)}</h1>
<p class="period" style="text-align:center"><strong>From:</strong> {_esc(period_start)} &nbsp;&mdash;&nbsp; <strong>To:</strong> {_esc(period_end)}</p>
<div class="header-rule"></div>
<div class="dashboard-section" style="margin-top:8px">
{cards_html}
</div>
<div class="gen-footer">
Generated by Hercules Reporting Module on {datetime.now().strftime('%d/%m/%Y, %H:%M:%S')}
</div>
</body></html>"""


def _generate_paginated_html(report_name, sections, tag_data, from_dt, to_dt):
    """Generate HTML report from paginated (Table Report) sections, styled to match frontend."""
    period_start = from_dt.strftime('%d/%m/%Y, %H:%M')
    period_end = to_dt.strftime('%d/%m/%Y, %H:%M')
    hercules_uri, asm_uri, client_logo_uri = _get_logo_data_uris()
    logo_html = _build_logo_header_html(hercules_uri, asm_uri, client_logo_uri)
    body_parts = [logo_html]

    total_rows = sum(len(s.get('rows', [])) for s in (sections or []) if s.get('type') == 'table')

    for s in (sections or []):
        s_type = s.get('type', '')

        if s_type == 'header':
            title = s.get('title', report_name) or report_name
            subtitle = s.get('subtitle', '')
            align = s.get('align', 'center')
            body_parts.append(f'<div style="text-align:{_esc(align)};margin-bottom:0">')
            body_parts.append(f'<h1 class="report-title">{_esc(title)}</h1>')
            if subtitle:
                body_parts.append(f'<p class="subtitle">{_esc(subtitle)}</p>')

            # Status field
            status_src = s.get('statusSourceType', 'static')
            status_val = None
            if status_src == 'static':
                sv = s.get('statusValue', '')
                if sv is not None and sv != '':
                    status_val = str(sv)
            elif status_src == 'tag' and s.get('statusTagName'):
                cell = {'sourceType': 'tag', 'tagName': s['statusTagName'],
                        'decimals': 0, 'unit': '', 'customUnit': ''}
                status_val = _resolve_cell(cell, tag_data)
            elif status_src == 'formula' and s.get('statusFormula'):
                cell = {'sourceType': 'formula', 'formula': s['statusFormula'],
                        'decimals': 0, 'unit': '', 'customUnit': ''}
                status_val = _resolve_cell(cell, tag_data)
            elif status_src == 'group' and s.get('statusGroupTags'):
                cell = {'sourceType': 'group', 'groupTags': s['statusGroupTags'],
                        'aggregation': s.get('statusAggregation', 'avg'),
                        'decimals': 0, 'unit': '', 'customUnit': ''}
                status_val = _resolve_cell(cell, tag_data)

            if status_val and status_val not in ('—', ''):
                status_label = s.get('statusLabel', 'Status')
                body_parts.append(f'<p class="subtitle">{_esc(status_label)}: {_esc(status_val)}</p>')

            if s.get('showDateRange', True):
                body_parts.append(f'<p class="period"><strong>From:</strong> {_esc(period_start)} &nbsp;&mdash;&nbsp; <strong>To:</strong> {_esc(period_end)}</p>')
            body_parts.append('<div class="header-rule"></div>')
            body_parts.append('</div>')

        elif s_type == 'kpi-row':
            label = s.get('label', '')
            kpis = s.get('kpis', [])
            if label:
                body_parts.append(f'<div class="kpi-section-label">{_esc(label)}</div>')
            kpi_items = []
            for k in kpis:
                kpi_label = k.get('label', '')
                cell_data = {
                    'sourceType': k.get('sourceType', 'tag'),
                    'tagName': k.get('tagName', ''),
                    'formula': k.get('formula', ''),
                    'unit': k.get('unit', ''),
                    'customUnit': k.get('customUnit', ''),
                    'decimals': k.get('decimals', 1),
                    'groupTags': k.get('groupTags', []),
                    'aggregation': k.get('aggregation', 'avg'),
                }
                val = _resolve_cell(cell_data, tag_data)
                kpi_items.append(
                    f'<td><span class="kpi-label">{_esc(kpi_label)}</span><br/>'
                    f'<span class="kpi-value">{_esc(val)}</span></td>'
                )
            if kpi_items:
                body_parts.append(f'<table class="kpi-row-table"><tr>{"".join(kpi_items)}</tr></table>')

        elif s_type == 'table':
            label = s.get('label', '')
            columns = s.get('columns', [])
            rows = s.get('rows', [])
            if label:
                body_parts.append(f'<div class="section-label">{_esc(label)}</div>')

            if not columns:
                continue

            # Column widths: <colgroup> + same width on each <td> (xhtml2pdf aligns thead/tbody reliably)
            num_cols = len(columns)
            default_w = f'{100.0 / num_cols:.1f}%' if num_cols else '100%'
            col_widths = []
            colgroup_parts = []
            header_cells = ''
            for c in columns:
                cw = c.get('width', '')
                if cw and str(cw).strip().lower() != 'auto':
                    w_attr = _esc(str(cw).strip())
                else:
                    w_attr = default_w
                col_widths.append(w_attr)
                colgroup_parts.append(f'<col width="{w_attr}" />')
                align = _esc(c.get('align', 'left'))
                header_cells += f'<th style="text-align:{align}">{_esc(c.get("header", ""))}</th>'
            cg = f'<colgroup>{"".join(colgroup_parts)}</colgroup>' if colgroup_parts else ''
            body_parts.append(
                f'<table class="data-table" width="100%">{cg}<thead><tr>{header_cells}</tr></thead><tbody>'
            )

            visible_row_idx = 0
            for row in rows:
                if _is_row_hidden(row, tag_data):
                    continue
                cells = row.get('cells', [])
                stripe = ' class="alt-row"' if visible_row_idx % 2 == 1 else ''
                td_parts = []
                for i in range(num_cols):
                    align = columns[i].get('align', 'left') if i < len(columns) else 'left'
                    w_style = col_widths[i] if i < len(col_widths) else default_w
                    if i < len(cells):
                        val = _resolve_cell(cells[i], tag_data)
                    else:
                        val = ''
                    td_parts.append(
                        f'<td style="text-align:{_esc(align)};width:{w_style}">{_esc(val)}</td>'
                    )
                body_parts.append(f'<tr{stripe}>{"".join(td_parts)}</tr>')
                visible_row_idx += 1

            # Summary row (per-column or legacy)
            has_summary = s.get('showSummaryRow') or any(
                c.get('summary', {}).get('enabled') or (c.get('summary', {}).get('type') and c.get('summary', {}).get('type') != 'none')
                for c in columns
            )
            if has_summary:
                has_per_col = any(
                    c.get('summary', {}).get('type') and c['summary']['type'] != 'none'
                    for c in columns
                )
                summary_cells = []
                for ci, col in enumerate(columns):
                    sm = col.get('summary', {})
                    sm_type = sm.get('type', 'none')
                    w_ci = col_widths[ci] if ci < len(col_widths) else default_w

                    if sm_type == 'label':
                        summary_cells.append(
                            f'<td class="summary-row" style="text-align:{_esc(col.get("align", "left"))};width:{w_ci}">'
                            f'{_esc(sm.get("label", s.get("summaryLabel", "Total")))}</td>'
                        )
                    elif sm_type == 'formula':
                        result = _evaluate_formula(sm.get('formula', ''), tag_data)
                        sm_unit = sm.get('unit', '')
                        # Derive decimals from first data cell in this column
                        col_dec = 0
                        for row in rows:
                            c0 = row.get('cells', [None] * (ci + 1))[ci] if ci < len(row.get('cells', [])) else None
                            if c0:
                                rd = c0.get('decimals')
                                col_dec = int(rd) if rd is not None and rd != '' else 0
                                break
                        val_str = f"{result:,.{col_dec}f}" if isinstance(result, (int, float)) and result is not None else '—'
                        if sm_unit and val_str != '—':
                            val_str = f"{val_str} {sm_unit}"
                        summary_cells.append(
                            f'<td class="summary-row" style="text-align:{_esc(col.get("align", "right"))};width:{w_ci}">'
                            f'{_esc(val_str)}</td>'
                        )
                    elif sm_type in ('sum', 'avg', 'min', 'max', 'count'):
                        # Aggregate from visible row values in this column
                        col_vals = []
                        col_dec = 0
                        for row in rows:
                            if _is_row_hidden(row, tag_data):
                                continue
                            cell = row.get('cells', [None] * (ci + 1))[ci] if ci < len(row.get('cells', [])) else None
                            if cell:
                                # Pick up decimals from first valid cell
                                if not col_vals:
                                    rd = cell.get('decimals')
                                    col_dec = int(rd) if rd is not None and rd != '' else 0
                                rv = _resolve_cell(cell, tag_data)
                                cleaned = re.sub(r'[^0-9.\-]', '', str(rv))
                                if cleaned:
                                    try:
                                        col_vals.append(float(cleaned))
                                    except (TypeError, ValueError):
                                        pass
                        if col_vals:
                            if sm_type == 'sum':
                                agg = sum(col_vals)
                            elif sm_type == 'avg':
                                agg = sum(col_vals) / len(col_vals)
                            elif sm_type == 'min':
                                agg = min(col_vals)
                            elif sm_type == 'max':
                                agg = max(col_vals)
                            elif sm_type == 'count':
                                agg = len(col_vals)
                            else:
                                agg = None
                            val_str = f"{agg:,.{col_dec}f}" if agg is not None else '—'
                            sm_unit = sm.get('unit', '')
                            if sm_unit and val_str != '—':
                                val_str = f"{val_str} {sm_unit}"
                            agg_label = f"{sm.get('label', '')}: " if sm.get('label') else ''
                            summary_cells.append(
                                f'<td class="summary-row" style="text-align:{_esc(col.get("align", "right"))};width:{w_ci}">'
                                f'{_esc(agg_label)}{_esc(val_str)}</td>'
                            )
                        else:
                            summary_cells.append(
                                f'<td class="summary-row" style="text-align:{_esc(col.get("align", "right"))};width:{w_ci}">—</td>'
                            )
                    elif not has_per_col:
                        # Legacy mode
                        if ci == 0:
                            colspan = max(1, len(columns) - 1)
                            w_span = w_ci if colspan <= 1 else ''
                            style = f'text-align:right;width:{w_span}' if w_span else 'text-align:right'
                            summary_cells.append(
                                f'<td class="summary-row" style="{style}" colspan="{colspan}">'
                                f'{_esc(s.get("summaryLabel", "Total"))}</td>'
                            )
                        elif ci == len(columns) - 1:
                            w_last = col_widths[-1] if col_widths else default_w
                            if s.get('summaryFormula'):
                                result = _evaluate_formula(s['summaryFormula'], tag_data)
                                su = s.get('summaryUnit', '')
                                # Derive decimals from first data cell in last column
                                col_dec = 0
                                for row in rows:
                                    c0 = row.get('cells', [])[-1] if row.get('cells') else None
                                    if c0:
                                        rd = c0.get('decimals')
                                        col_dec = int(rd) if rd is not None and rd != '' else 0
                                        break
                                val_str = f"{result:,.{col_dec}f}" if isinstance(result, (int, float)) and result is not None else '—'
                                if su and val_str != '—':
                                    val_str = f"{val_str} {su}"
                                summary_cells.append(
                                    f'<td class="summary-row" style="text-align:right;width:{w_last}">{_esc(val_str)}</td>'
                                )
                            else:
                                summary_cells.append(
                                    f'<td class="summary-row" style="text-align:right;width:{w_last}">—</td>'
                                )
                        # else: skip (covered by colspan)
                    else:
                        # Per-column mode but this column has no summary
                        if ci == 0:
                            summary_cells.append(
                                f'<td class="summary-row" style="text-align:left;width:{w_ci}">'
                                f'{_esc(s.get("summaryLabel", "Total"))}</td>'
                            )
                        else:
                            summary_cells.append(
                                f'<td class="summary-row" style="width:{w_ci}"></td>'
                            )

                body_parts.append(f'<tr class="summary-row">{"".join(summary_cells)}</tr>')

            body_parts.append('</tbody></table>')

        elif s_type == 'text-block':
            content = s.get('content', '')
            fs = s.get('fontSize', '14px')
            fw = s.get('fontWeight', '600')
            align = s.get('align', 'left')
            color = s.get('color', '#0f172a')
            body_parts.append(
                f'<div class="text-block" style="font-size:{_esc(fs)};font-weight:{_esc(fw)};'
                f'text-align:{_esc(align)};color:{_esc(color)}">{_esc(content)}</div>'
            )

        elif s_type == 'spacer':
            h = s.get('height', 16)
            body_parts.append(f'<div style="height:{h}px"></div>')

        elif s_type == 'signature-block':
            fields = s.get('fields', [])
            if fields:
                body_parts.append('<div class="sig-block"><table><tr>')
                for f in fields:
                    body_parts.append(
                        f'<td><div class="sig-label">{_esc(f.get("label", ""))}</div>'
                        f'<div class="sig-line">{_esc(f.get("value", ""))}&nbsp;</div>'
                        f'<div class="sig-date">Date: _______________</div></td>'
                    )
                body_parts.append('</tr></table></div>')

    content_html = '\n'.join(body_parts)

    footer_records = f'Records: {total_rows}' if total_rows > 0 else ''
    gen_ts = datetime.now().strftime('%d/%m/%Y, %H:%M:%S')
    footer_tbl = (
        f'<table class="gen-footer-table" width="100%" cellpadding="0" cellspacing="0"><tr>'
        f'<td style="text-align:left;width:35%">{_esc(footer_records)}</td>'
        f'<td style="text-align:right;width:65%">Generated by Hercules Reporting Module on {_esc(gen_ts)}</td>'
        f'</tr></table>'
    )
    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><style>{_SHARED_CSS}</style></head>
<body style="padding: 3mm 10mm 5mm 10mm; width: 190mm;">
{content_html}
{footer_tbl}
</body></html>"""


def _generate_report_html(report_name, layout_config, tag_data, from_dt, to_dt):
    """Route to the correct renderer based on report type."""
    report_type = layout_config.get('reportType', 'dashboard')
    paginated_sections = layout_config.get('paginatedSections', [])

    if report_type == 'paginated' and paginated_sections:
        return _generate_paginated_html(report_name, paginated_sections, tag_data, from_dt, to_dt)
    else:
        return _generate_dashboard_html(report_name, layout_config, tag_data, from_dt, to_dt)


# ── PDF conversion ───────────────────────────────────────────────────────────

def _html_to_pdf(html_content):
    """Convert HTML string to PDF bytes using xhtml2pdf."""
    from xhtml2pdf import pisa
    from io import BytesIO
    buf = BytesIO()
    result = pisa.CreatePDF(html_content, dest=buf)
    if result.err:
        raise RuntimeError(f"PDF generation failed with {result.err} error(s)")
    return buf.getvalue()


# ── Public API for report export ──────────────────────────────────────────────

def generate_report_xlsx(report_name, layout_config, from_dt, to_dt):
    """Public wrapper: generate an Excel report from a template config and date range.
    Returns: bytes (xlsx file content)
    """
    tag_names = extract_all_tags(layout_config)
    tag_data = _fetch_tag_data_multi_agg(layout_config, tag_names, from_dt, to_dt)
    return _generate_xlsx(report_name, layout_config, tag_data, from_dt, to_dt)


# ── Excel (XLSX) generation ──────────────────────────────────────────────────

def _generate_xlsx(report_name, layout_config, tag_data, from_dt, to_dt):
    """Generate an Excel workbook from a report template and tag data.
    Returns: bytes (xlsx file content)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = Workbook()
    wb.properties.title = report_name
    wb.properties.creator = "Hercules Reporting Module"

    # Shared styles
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='1A2233', end_color='1A2233', fill_type='solid')
    alt_fill = PatternFill(start_color='F5F8FB', end_color='F5F8FB', fill_type='solid')
    summary_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')
    summary_font = Font(name='Calibri', bold=True, size=10)
    title_font = Font(name='Calibri', bold=True, size=14)
    subtitle_font = Font(name='Calibri', size=11, color='666666')
    kpi_label_font = Font(name='Calibri', size=9, color='888888')
    kpi_value_font = Font(name='Calibri', bold=True, size=12)
    thin_border = Border(
        bottom=Side(style='thin', color='E3E9F0')
    )
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    num_align = Alignment(horizontal='right', vertical='center')
    text_align = Alignment(horizontal='left', vertical='center')

    report_type = layout_config.get('reportType', 'paginated')

    try:
        if report_type == 'paginated':
            _xlsx_paginated(wb, layout_config, tag_data, from_dt, to_dt, report_name,
                            header_font, header_fill, alt_fill, summary_fill, summary_font,
                            title_font, subtitle_font, kpi_label_font, kpi_value_font,
                            thin_border, header_align, num_align, text_align)
        else:
            _xlsx_dashboard(wb, layout_config, tag_data, from_dt, to_dt, report_name,
                            header_font, header_fill, alt_fill, title_font, subtitle_font,
                            kpi_label_font, kpi_value_font, thin_border, num_align, text_align)
    except Exception as e:
        logger.error(f"Excel generation failed for '{report_name}': {e}", exc_info=True)
        raise RuntimeError(f"Excel generation failed: {e}")

    # Auto-fit column widths on all sheets
    for ws in wb.worksheets:
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    cell_len = len(str(cell.value)) if cell.value is not None else 0
                    max_len = max(max_len, cell_len)
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_paginated(wb, layout_config, tag_data, from_dt, to_dt, report_name,
                    header_font, header_fill, alt_fill, summary_fill, summary_font,
                    title_font, subtitle_font, kpi_label_font, kpi_value_font,
                    thin_border, header_align, num_align, text_align):
    """Render paginated (table) report sections to Excel sheets."""
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font

    ws = wb.active
    ws.title = report_name[:31]  # Excel sheet name limit
    row_idx = 1

    sections = layout_config.get('paginatedSections', []) or layout_config.get('sections', [])
    period = f"Start: {from_dt.strftime('%d/%m/%Y, %H:%M')}  |  End: {to_dt.strftime('%d/%m/%Y, %H:%M')}" if from_dt and to_dt else ''

    for section in sections:
        stype = section.get('type', '')

        # ── Header section ──
        if stype == 'header':
            title = section.get('title', report_name)
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
            ws.cell(row=row_idx, column=1, value=title).font = title_font
            row_idx += 1

            subtitle = section.get('subtitle', '')
            if subtitle:
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
                ws.cell(row=row_idx, column=1, value=subtitle).font = subtitle_font
                row_idx += 1

            if period and section.get('showDateRange', True):
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
                ws.cell(row=row_idx, column=1, value=period).font = subtitle_font
                row_idx += 1

            # Status value if configured
            status_src = section.get('statusSourceType')
            if status_src:
                status_cell = {
                    'sourceType': status_src,
                    'tagName': section.get('statusTagName', ''),
                    'formula': section.get('statusFormula', ''),
                    'groupTags': section.get('statusGroupTags', []),
                    'aggregation': section.get('statusAggregation', 'avg'),
                    'value': section.get('statusValue', ''),
                    'decimals': 1, 'unit': '',
                }
                val, _, _ = _resolve_cell_raw(status_cell, tag_data)
                label = section.get('statusLabel', 'Status')
                ws.cell(row=row_idx, column=1, value=label).font = kpi_label_font
                c = ws.cell(row=row_idx, column=2, value=val)
                if isinstance(val, (int, float)):
                    c.alignment = num_align
                row_idx += 1

            row_idx += 1  # Blank row after header

        # ── KPI row section ──
        elif stype == 'kpi-row':
            kpis = section.get('kpis', [])
            for col_i, kpi in enumerate(kpis):
                col = col_i * 2 + 1
                ws.cell(row=row_idx, column=col, value=kpi.get('label', '')).font = kpi_label_font
                val, unit, dec = _resolve_cell_raw(kpi, tag_data)
                c = ws.cell(row=row_idx, column=col + 1, value=val)
                if isinstance(val, (int, float)):
                    c.number_format = f'#,##0.{"0" * dec}'
                    c.alignment = num_align
                c.font = kpi_value_font
            row_idx += 2  # Blank row after KPIs

        # ── Table section ──
        elif stype == 'table':
            label = section.get('label', '')
            if label:
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max(len(section.get('columns', [])), 1))
                ws.cell(row=row_idx, column=1, value=label).font = Font(name='Calibri', bold=True, size=11)
                row_idx += 1

            columns = section.get('columns', [])
            rows = section.get('rows', [])
            num_cols = len(columns)

            if num_cols == 0:
                continue

            # Column headers
            header_row = row_idx
            for ci, col in enumerate(columns):
                hdr_text = col.get('header', f'Col {ci+1}')
                # Append unit to header if commonly used
                c = ws.cell(row=row_idx, column=ci + 1, value=hdr_text)
                c.font = header_font
                c.fill = header_fill
                c.alignment = header_align
                c.border = thin_border
            row_idx += 1

            # Freeze header row
            ws.freeze_panes = ws.cell(row=row_idx, column=1)

            # Data rows
            data_start_row = row_idx
            visible_row_count = 0
            for row in rows:
                if _is_row_hidden(row, tag_data):
                    continue
                cells = row.get('cells', [])
                for ci in range(num_cols):
                    cell_def = cells[ci] if ci < len(cells) else None
                    val, unit, dec = _resolve_cell_raw(cell_def, tag_data)
                    c = ws.cell(row=row_idx, column=ci + 1, value=val)
                    if isinstance(val, (int, float)):
                        c.number_format = f'#,##0.{"0" * dec}'
                        c.alignment = num_align
                    else:
                        c.alignment = text_align
                    c.border = thin_border
                    # Alternating row color
                    if visible_row_count % 2 == 1:
                        c.fill = alt_fill
                visible_row_count += 1
                row_idx += 1
            data_end_row = row_idx - 1

            # Summary row with native Excel formulas
            if section.get('showSummaryRow') and visible_row_count > 0:
                for ci, col in enumerate(columns):
                    sm = col.get('summary', {})
                    sm_type = sm.get('type', 'none') if sm else 'none'
                    col_letter = get_column_letter(ci + 1)
                    c = ws.cell(row=row_idx, column=ci + 1)
                    c.font = summary_font
                    c.fill = summary_fill
                    c.border = thin_border

                    if sm_type == 'label':
                        c.value = sm.get('label', '')
                        c.alignment = text_align
                    elif sm_type == 'sum':
                        c.value = f'=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})'
                        c.alignment = num_align
                    elif sm_type == 'avg':
                        c.value = f'=AVERAGE({col_letter}{data_start_row}:{col_letter}{data_end_row})'
                        c.alignment = num_align
                    elif sm_type == 'min':
                        c.value = f'=MIN({col_letter}{data_start_row}:{col_letter}{data_end_row})'
                        c.alignment = num_align
                    elif sm_type == 'max':
                        c.value = f'=MAX({col_letter}{data_start_row}:{col_letter}{data_end_row})'
                        c.alignment = num_align
                    elif sm_type == 'count':
                        c.value = f'=COUNTA({col_letter}{data_start_row}:{col_letter}{data_end_row})'
                        c.alignment = num_align
                    elif sm_type == 'formula':
                        # Can't express as Excel formula — pre-compute
                        formula_str = sm.get('formula', '')
                        result = _evaluate_formula(formula_str, tag_data)
                        c.value = result
                        c.alignment = num_align
                    # else: leave empty

                row_idx += 1

            # Auto-filter on data columns
            if visible_row_count > 0:
                ws.auto_filter.ref = f"A{header_row}:{get_column_letter(num_cols)}{data_end_row}"

            row_idx += 1  # Blank row between tables

        # Skip signature, text, spacer sections
        elif stype in ('signature', 'text', 'spacer'):
            continue


def _xlsx_dashboard(wb, layout_config, tag_data, from_dt, to_dt, report_name,
                    header_font, header_fill, alt_fill, title_font, subtitle_font,
                    kpi_label_font, kpi_value_font, thin_border, num_align, text_align):
    """Render dashboard report widgets to an Excel sheet."""
    from openpyxl.styles import Font

    ws = wb.active
    ws.title = "Dashboard Summary"
    row_idx = 1

    period = f"Start: {from_dt.strftime('%d/%m/%Y, %H:%M')}  |  End: {to_dt.strftime('%d/%m/%Y, %H:%M')}" if from_dt and to_dt else ''

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.cell(row=1, column=1, value=report_name).font = title_font
    row_idx = 2
    if period:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
        ws.cell(row=2, column=1, value=period).font = subtitle_font
        row_idx = 3
    row_idx += 1  # Blank row

    section_title_font = Font(name='Calibri', bold=True, size=11)
    widget_group_font = Font(name='Calibri', bold=True, size=10)

    for section_label, flat_widgets in _dashboard_sections_for_distribution(layout_config):
        if section_label:
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
            ws.cell(row=row_idx, column=1, value=section_label).font = section_title_font
            row_idx += 2

        for widget in flat_widgets:
            wtype = widget.get('type', '')
            config = widget.get('config', {}) or {}
            label = config.get('title', '') or config.get('label', '') or widget.get('name', '')

            if wtype in ('kpi', 'gauge', 'stat', 'progress', 'sparkline'):
                ds = config.get('dataSource') or {}
                raw = _resolve_widget_datasource_value(ds, tag_data)
                ws.cell(row=row_idx, column=1, value=label).font = kpi_label_font
                c = ws.cell(row=row_idx, column=2)
                try:
                    c.value = float(raw) if raw is not None else None
                    c.number_format = '#,##0.00'
                    c.alignment = num_align
                except (TypeError, ValueError):
                    c.value = str(raw) if raw is not None and raw != '' else '—'
                c.font = kpi_value_font
                unit = config.get('unit', '')
                if unit:
                    ws.cell(row=row_idx, column=3, value=unit).font = kpi_label_font
                row_idx += 1

            elif wtype == 'silo':
                ws.cell(row=row_idx, column=1, value=label).font = widget_group_font
                row_idx += 1
                for sub_label, sub_key in [('Level', 'tagName'), ('Capacity', 'capacityTag'), ('Tons', 'tonsTag')]:
                    tag_name = (config.get('dataSource') or {}).get(sub_key, '') if sub_key == 'tagName' else config.get(sub_key, '')
                    raw = tag_data.get(tag_name)
                    ws.cell(row=row_idx, column=1, value=f'  {sub_label}').font = kpi_label_font
                    c = ws.cell(row=row_idx, column=2)
                    try:
                        c.value = float(raw) if raw is not None else None
                        c.number_format = '#,##0.0'
                        c.alignment = num_align
                    except (TypeError, ValueError):
                        c.value = str(raw) if raw else '—'
                    row_idx += 1
                row_idx += 1

            elif wtype in ('chart', 'barchart', 'piechart'):
                ws.cell(row=row_idx, column=1, value=label).font = widget_group_font
                row_idx += 1
                for series in _chart_series_list_for_export(config):
                    s_label = series.get('label', '')
                    tag_name = (series.get('dataSource') or {}).get('tagName', '') or series.get('tagName', '')
                    raw = tag_data.get(tag_name) if tag_name else None
                    ws.cell(row=row_idx, column=1, value=f'  {s_label}').font = kpi_label_font
                    c = ws.cell(row=row_idx, column=2)
                    try:
                        c.value = float(raw) if raw is not None else None
                        c.number_format = '#,##0.00'
                        c.alignment = num_align
                    except (TypeError, ValueError):
                        c.value = str(raw) if raw else '—'
                    row_idx += 1
                row_idx += 1

            elif wtype == 'table':
                ws.cell(row=row_idx, column=1, value=label).font = widget_group_font
                row_idx += 1
                table_cols = config.get('tableColumns') or config.get('columns') or []
                for ci, col in enumerate(table_cols):
                    c = ws.cell(row=row_idx, column=ci + 1, value=col.get('label', ''))
                    c.font = header_font
                    c.fill = header_fill
                    c.border = thin_border
                row_idx += 1
                for ci, col in enumerate(table_cols):
                    src = col.get('sourceType', 'tag')
                    cell_def = {'sourceType': src, 'tagName': col.get('tagName', ''),
                                'formula': col.get('formula', ''), 'groupTags': col.get('groupTags', []),
                                'aggregation': col.get('aggregation', 'avg'), 'decimals': 2, 'unit': ''}
                    val, unit, dec = _resolve_cell_raw(cell_def, tag_data)
                    c = ws.cell(row=row_idx, column=ci + 1, value=val)
                    if isinstance(val, (int, float)):
                        c.number_format = f'#,##0.{"0" * dec}'
                        c.alignment = num_align
                    c.border = thin_border
                row_idx += 2

            elif wtype in ('text', 'header', 'image', 'spacer', 'divider'):
                continue

            else:
                ds = config.get('dataSource') or {}
                raw = _resolve_widget_datasource_value(ds, tag_data)
                if raw is None:
                    continue
                ws.cell(row=row_idx, column=1, value=label).font = kpi_label_font
                c = ws.cell(row=row_idx, column=2)
                try:
                    c.value = float(raw) if raw is not None else None
                    c.number_format = '#,##0.00'
                    c.alignment = num_align
                except (TypeError, ValueError):
                    c.value = str(raw)
                c.font = kpi_value_font
                unit = config.get('unit', '')
                if unit:
                    ws.cell(row=row_idx, column=3, value=unit).font = kpi_label_font
                row_idx += 1


# ── Email delivery ───────────────────────────────────────────────────────────

def _build_email_html(report_name, from_dt, to_dt, filename):
    """Build a professionally formatted HTML email body with logos and report info."""
    hercules_uri, asm_uri, client_logo_uri = _get_logo_data_uris()
    period_from = from_dt.strftime('%d/%m/%Y, %H:%M')
    period_to = to_dt.strftime('%d/%m/%Y, %H:%M')
    generated = datetime.now().strftime('%d/%m/%Y, %H:%M')

    # Logo images for the header (explicit width+height for Outlook compatibility)
    hercules_img = f'<img src="{hercules_uri}" alt="Hercules" width="auto" height="36" style="height:36px;width:auto;display:block" />' if hercules_uri else ''
    asm_img = f'<img src="{asm_uri}" alt="ASM" width="auto" height="32" style="height:32px;width:auto;display:inline-block;vertical-align:middle" />' if asm_uri else ''
    client_img = f'<img src="{client_logo_uri}" alt="" width="auto" height="32" style="height:32px;width:auto;max-width:100px;display:inline-block;vertical-align:middle" />' if client_logo_uri else ''
    right_logos = ''
    if client_img or asm_img:
        spacer = '&nbsp;&nbsp;' if client_img and asm_img else ''
        right_logos = f'{client_img}{spacer}{asm_img}'

    return f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background:#f1f5f9;font-family:Inter,system-ui,-apple-system,sans-serif;color:#1a1a2e">
<table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#f1f5f9;padding:24px 0">
<tr><td align="center">
<table role="presentation" width="600" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:8px;overflow:hidden;box-shadow:0 1px 3px rgba(0,0,0,0.08)">

  <!-- Logo header bar -->
  <tr>
    <td style="background:linear-gradient(135deg,#0f1b2d 0%,#1a3a5c 100%);padding:16px 24px">
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0">
        <tr>
          <td style="text-align:left">{hercules_img}</td>
          <td style="text-align:right">{right_logos}</td>
        </tr>
      </table>
    </td>
  </tr>

  <!-- Report title & period -->
  <tr>
    <td style="padding:28px 32px 0 32px;text-align:center">
      <div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:0.1em;color:#94a3b8;margin-bottom:6px">Scheduled Report</div>
      <div style="font-size:22px;font-weight:700;color:#0f172a;letter-spacing:-0.02em;margin-bottom:4px">{_esc(report_name)}</div>
      <div style="height:2px;width:60px;background:linear-gradient(90deg,#0f3460,#1a5276);margin:8px auto 12px auto;border-radius:1px"></div>
    </td>
  </tr>

  <!-- Details card -->
  <tr>
    <td style="padding:0 32px 24px 32px">
      <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:6px">
        <tr>
          <td style="padding:14px 20px;border-bottom:1px solid #e2e8f0">
            <div style="font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:0.06em;color:#94a3b8;margin-bottom:2px">Report Period</div>
            <div style="font-size:14px;font-weight:600;color:#334155"><strong>From:</strong> {_esc(period_from)} &nbsp;&mdash;&nbsp; <strong>To:</strong> {_esc(period_to)}</div>
          </td>
        </tr>
        <tr>
          <td style="padding:14px 20px;border-bottom:1px solid #e2e8f0">
            <div style="font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:0.06em;color:#94a3b8;margin-bottom:2px">Generated</div>
            <div style="font-size:14px;font-weight:600;color:#334155">{_esc(generated)}</div>
          </td>
        </tr>
        <tr>
          <td style="padding:14px 20px">
            <div style="font-size:10px;font-weight:600;text-transform:uppercase;letter-spacing:0.06em;color:#94a3b8;margin-bottom:2px">Attachment</div>
            <div style="font-size:14px;font-weight:600;color:#334155">{_esc(filename)}</div>
          </td>
        </tr>
      </table>
    </td>
  </tr>

  <!-- Message -->
  <tr>
    <td style="padding:0 32px 28px 32px">
      <p style="font-size:13px;color:#475569;line-height:1.6;margin:0">
        Please find the scheduled report <strong>{_esc(report_name)}</strong> attached to this email.
        The attachment uses the same layout as the dashboard viewer (including every tab when the report has multiple dashboard tabs).
        The report covers the period shown above. If you have questions about this report, please
        contact your system administrator.
      </p>
    </td>
  </tr>

  <!-- Footer -->
  <tr>
    <td style="background:#f8fafc;border-top:1px solid #e2e8f0;padding:14px 32px;text-align:center">
      <div style="font-size:11px;color:#94a3b8;line-height:1.5">
        This is an automated email from the <strong style="color:#64748b">Hercules Reporting Module</strong>.
      </div>
    </td>
  </tr>

</table>
</td></tr>
</table>
</body></html>"""


def _send_email(recipients, subject, body_html, attachments=None):
    """Send HTML email with optional attachments using configured method (Resend or SMTP).

    Args:
        attachments: list of (filename, bytes) tuples, or None
    """
    from smtp_config import get_smtp_config
    cfg = get_smtp_config()

    # ── Resend (default) — falls through to SMTP if the package is missing ──
    if cfg.get('send_method', 'resend') == 'resend':
        try:
            from smtp_config import send_email_resend
            return send_email_resend(recipients, subject, body_html, attachments=attachments)
        except ImportError:
            logger.warning("'resend' package not installed, falling back to SMTP")

    # ── SMTP fallback ──
    if not cfg.get('smtp_server'):
        return {'success': False, 'error': 'No SMTP server configured'}

    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = cfg.get('from_address') or cfg.get('username', '')
    msg['To'] = ', '.join(recipients)

    # Plain-text fallback
    plain_text = "Please find the attached report(s).\n"
    msg.set_content(plain_text)
    msg.add_alternative(body_html, subtype='html')

    if attachments:
        import mimetypes
        for filename, content_bytes in attachments:
            mime_type, _ = mimetypes.guess_type(filename)
            if mime_type:
                maintype, subtype = mime_type.split('/', 1)
            else:
                maintype, subtype = 'application', 'octet-stream'
            msg.add_attachment(
                content_bytes,
                maintype=maintype,
                subtype=subtype,
                filename=filename,
            )

    port = cfg.get('smtp_port', 465)
    try:
        if port == 465:
            with smtplib.SMTP_SSL(cfg['smtp_server'], port, timeout=30) as server:
                server.login(cfg['username'], cfg['password'])
                server.send_message(msg)
        else:
            with smtplib.SMTP(cfg['smtp_server'], port, timeout=30) as server:
                server.starttls()
                server.login(cfg['username'], cfg['password'])
                server.send_message(msg)
        return {'success': True}
    except Exception as e:
        return {'success': False, 'error': str(e)}


# ── Disk delivery ────────────────────────────────────────────────────────────

def _save_to_disk(save_path, filename, content_bytes):
    """Save report file to disk at the user-specified path."""
    if not save_path or not save_path.strip():
        save_path = DEFAULT_SAVE_DIR
    resolved = os.path.realpath(os.path.join(save_path, filename))
    os.makedirs(os.path.dirname(resolved), exist_ok=True)
    with open(resolved, 'wb') as f:
        f.write(content_bytes)
    return resolved


# ── Main execution ───────────────────────────────────────────────────────────

def execute_distribution_rule(rule_id):
    """
    Execute a single distribution rule: load report, fetch data, generate
    PDF/HTML, deliver via email/disk, and update run status in DB.
    """
    get_conn = _get_db_connection()

    # 1. Load rule from DB
    with closing(get_conn()) as conn:
        actual_conn = conn._conn if hasattr(conn, '_conn') else conn
        cur = actual_conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM distribution_rules WHERE id = %s", (rule_id,))
        rule = cur.fetchone()

    if not rule:
        return {'success': False, 'error': f'Rule {rule_id} not found'}

    rule = dict(rule)

    # Support multi-report: report_ids array, fallback to [report_id]
    report_ids = rule.get('report_ids')
    if isinstance(report_ids, str):
        report_ids = json.loads(report_ids)
    if not report_ids or report_ids == []:
        rid = rule.get('report_id')
        report_ids = [rid] if rid and rid > 0 else []

    if not report_ids:
        return {'success': False, 'error': 'No reports configured for this rule'}

    try:
        # 2. Determine time range (shared across all reports)
        from_dt, to_dt = _time_range_for_schedule(rule['schedule_type'])
        fmt = rule.get('format', 'pdf')
        ext = {'pdf': 'pdf', 'html': 'html', 'xlsx': 'xlsx'}.get(fmt, 'pdf')
        timestamp_str = datetime.now().strftime('%Y%m%d_%H%M')

        # 3. Generate each report
        attachments = []  # list of (filename, content_bytes)
        report_names = []
        skipped = []

        for rid in report_ids:
            with closing(get_conn()) as conn:
                actual_conn = conn._conn if hasattr(conn, '_conn') else conn
                cur = actual_conn.cursor(cursor_factory=RealDictCursor)
                cur.execute(
                    "SELECT id, name, layout_config FROM report_builder_templates WHERE id = %s",
                    (rid,)
                )
                template = cur.fetchone()

            if not template:
                skipped.append(f"Report {rid} not found (deleted?)")
                logger.warning(f"Distribution rule {rule_id}: report {rid} not found, skipping")
                continue

            template = dict(template)
            report_name = template['name']
            report_names.append(report_name)
            layout_config = template['layout_config']
            if isinstance(layout_config, str):
                layout_config = json.loads(layout_config)

            tag_names = extract_all_tags(layout_config)
            tag_data = _fetch_tag_data_multi_agg(layout_config, tag_names, from_dt, to_dt)
            if fmt == 'xlsx':
                content_bytes = _generate_xlsx(report_name, layout_config, tag_data, from_dt, to_dt)
            else:
                html_content = _generate_report_html(report_name, layout_config, tag_data, from_dt, to_dt)
                if fmt == 'pdf':
                    content_bytes = _html_to_pdf(html_content)
                else:
                    content_bytes = html_content.encode('utf-8')

            safe_name = re.sub(r'[^\w\-]', '_', report_name)
            filename = f"{safe_name}_{timestamp_str}.{ext}"
            attachments.append((filename, content_bytes))

        if not attachments:
            raise ValueError(f"All reports missing: {'; '.join(skipped)}")

        # 4. Deliver
        delivery = rule['delivery_method']
        recipients = rule.get('recipients', [])
        if isinstance(recipients, str):
            recipients = json.loads(recipients)

        errors = list(skipped)  # include skipped reports as warnings

        if delivery in ('email', 'both') and recipients:
            names_str = ', '.join(report_names)
            subject = f"Hercules Report: {names_str} — {datetime.now().strftime('%Y-%m-%d')}"
            email_html = _build_email_html(names_str, from_dt, to_dt,
                                           ', '.join(fn for fn, _ in attachments))

            # ── AI Summary injection (Phase 1) ──────────────────────────
            if rule.get('include_ai_summary'):
                try:
                    all_tag_data = {}
                    all_layout_configs = {}
                    for rid in report_ids:
                        with closing(get_conn()) as conn2:
                            actual2 = conn2._conn if hasattr(conn2, '_conn') else conn2
                            cur2 = actual2.cursor(cursor_factory=RealDictCursor)
                            cur2.execute("SELECT name, layout_config FROM report_builder_templates WHERE id = %s", (rid,))
                            tpl = cur2.fetchone()
                            if tpl:
                                lc = tpl['layout_config']
                                if isinstance(lc, str):
                                    lc = json.loads(lc)
                                tags = extract_all_tags(lc)
                                td = _fetch_tag_data_multi_agg(lc, tags, from_dt, to_dt)
                                all_tag_data.update(td)
                                all_layout_configs[tpl['name'] or f'Report {rid}'] = lc

                    summary = _generate_ai_summary(
                        report_names=report_names,
                        tag_data=all_tag_data,
                        from_dt=from_dt,
                        to_dt=to_dt,
                        layout_configs=all_layout_configs,
                    )
                    if summary:
                        email_html = _prepend_summary_to_email(summary, email_html)
                except Exception as e:
                    logger.warning("AI summary generation failed, sending without: %s", e)
            # ── End AI Summary ───────────────────────────────────────────

            email_result = _send_email(recipients, subject, email_html, attachments=attachments)
            if not email_result['success']:
                errors.append(f"Email: {email_result['error']}")

        if delivery in ('disk', 'both') and rule.get('save_path'):
            for filename, content_bytes in attachments:
                try:
                    saved_path = _save_to_disk(rule['save_path'], filename, content_bytes)
                    logger.info(f"Report saved to {saved_path}")
                except Exception as e:
                    errors.append(f"Disk ({filename}): {str(e)}")

        # 5. Update run status
        real_errors = [e for e in errors if not e.startswith('Report ') or 'not found' not in e]
        success = len(real_errors) == 0
        status = 'success' if success else 'failed'
        error_msg = '; '.join(errors) if errors else None

        file_names_list = [fn for fn, _ in attachments] if attachments else []

        with closing(get_conn()) as conn:
            actual_conn = conn._conn if hasattr(conn, '_conn') else conn
            cur = actual_conn.cursor()
            # Log execution to audit trail
            try:
                cur.execute("""
                    INSERT INTO report_execution_log
                        (rule_id, report_ids, executed_at, time_range_from, time_range_to,
                         format, delivery_method, recipients, status, error_message, file_names)
                    VALUES (%s, %s::jsonb, NOW(), %s, %s, %s, %s, %s::jsonb, %s, %s, %s::jsonb)
                """, (
                    rule_id, json.dumps(rule.get('report_ids', [])),
                    from_dt, to_dt, fmt, delivery,
                    json.dumps(rule.get('recipients', [])),
                    status, error_msg, json.dumps(file_names_list),
                ))
            except Exception as log_err:
                logger.warning(f"Failed to log execution: {log_err}")
            cur.execute("""
                UPDATE distribution_rules
                SET last_run_at = NOW(), last_run_status = %s, last_run_error = %s
                WHERE id = %s
            """, (status, error_msg, rule_id))
            actual_conn.commit()

        if success:
            return {'success': True, 'message': f'{len(attachments)} report(s) delivered via {delivery}'}
        else:
            return {'success': False, 'error': '; '.join(errors)}

    except Exception as e:
        logger.error(f"Distribution rule {rule_id} failed: {e}", exc_info=True)
        # Update status to failed + log failure
        try:
            with closing(get_conn()) as conn:
                actual_conn = conn._conn if hasattr(conn, '_conn') else conn
                cur = actual_conn.cursor()
                try:
                    cur.execute("""
                        INSERT INTO report_execution_log
                            (rule_id, executed_at, status, error_message)
                        VALUES (%s, NOW(), 'failed', %s)
                    """, (rule_id, str(e)))
                except Exception:
                    pass
                cur.execute("""
                    UPDATE distribution_rules
                    SET last_run_at = NOW(), last_run_status = 'failed', last_run_error = %s
                    WHERE id = %s
                """, (str(e), rule_id))
                actual_conn.commit()
        except Exception:
            pass
        return {'success': False, 'error': str(e)}


# ── AI Summary helpers (Phase 1) ─────────────────────────────────────────────

# Daily rate limit for AI calls
_ai_call_count = 0
_ai_call_date = None
_AI_DAILY_CAP = 200


def _extract_report_context(layout_configs):
    """Build a human-readable description of report structures for the AI prompt.
    layout_configs: dict {report_name: layout_config_dict}
    """
    parts = []
    for name, lc in (layout_configs or {}).items():
        if not isinstance(lc, dict):
            continue
        report_type = lc.get('reportType', 'dashboard')

        # Paginated (table) reports
        paginated = lc.get('paginatedSections', [])
        if report_type == 'paginated' and paginated:
            lines = [f'Report "{name}" (Table Report):']
            for s in paginated:
                stype = s.get('type', '')
                if stype == 'table':
                    cols = s.get('columns', [])
                    col_headers = [c.get('header', '?') for c in cols]
                    label = s.get('label', 'Table')
                    lines.append(f'  Table "{label}": columns {col_headers}')
                    # Describe aggregation context per column from first row
                    rows = s.get('rows', [])
                    if rows:
                        for ci, cell in enumerate((rows[0].get('cells') or [])):
                            agg = cell.get('aggregation', '')
                            src = cell.get('sourceType', '')
                            unit_val = cell.get('unit', '')
                            is_cb = unit_val == '__checkbox__'
                            if ci < len(col_headers):
                                col_name = col_headers[ci]
                            else:
                                col_name = f'Col {ci}'
                            if src == 'static':
                                lines.append(f'    - "{col_name}": static label (row identifier)')
                            elif agg == 'delta':
                                lines.append(f'    - "{col_name}": delta aggregation (= amount produced/consumed in the period)')
                            elif agg == 'first':
                                lines.append(f'    - "{col_name}": first aggregation (= reading at start of period)')
                            elif agg in ('avg', 'sum', 'min', 'max', 'count'):
                                lines.append(f'    - "{col_name}": {agg} aggregation')
                            elif is_cb:
                                lines.append(f'    - "{col_name}": boolean status (on/off)')
                            elif agg == '' or agg == 'last':
                                lines.append(f'    - "{col_name}": last value (current/end reading)')
                        lines.append(f'    {len(rows)} data rows')
                    # Summary row info
                    has_summary = any(c.get('summary', {}).get('type', 'none') != 'none' for c in cols)
                    if has_summary:
                        sm_parts = []
                        for c in cols:
                            st = c.get('summary', {}).get('type', 'none')
                            if st != 'none':
                                sm_parts.append(f'{c.get("header","?")}={st}')
                        lines.append(f'    Summary row: {", ".join(sm_parts)}')
                elif stype == 'kpi-row':
                    kpis = s.get('kpis', [])
                    kpi_labels = [k.get('label', '?') for k in kpis]
                    lines.append(f'  KPI Row: {kpi_labels}')
            parts.append('\n'.join(lines))
            continue

        # Dashboard reports
        dt = lc.get('dashboardTabs', {})
        tabs = dt.get('tabs', []) if isinstance(dt.get('tabs'), list) else []
        if dt.get('enabled') and tabs:
            lines = [f'Report "{name}" (Dashboard):']
            for tab in tabs:
                tab_label = tab.get('label', '?')
                widgets = tab.get('widgets', [])
                widget_types = {}
                for w in widgets:
                    wt = w.get('type', '?')
                    cfg = w.get('config', {}) or {}
                    title = cfg.get('title') or cfg.get('label') or cfg.get('content', '')
                    if wt in ('kpi', 'gauge', 'stat', 'progress', 'sparkline'):
                        widget_types.setdefault('KPIs', []).append(title)
                    elif wt in ('chart', 'barchart', 'piechart'):
                        widget_types.setdefault('Charts', []).append(title)
                    elif wt == 'table':
                        col_labels = [c.get('label', '') for c in (cfg.get('tableColumns') or [])]
                        widget_types.setdefault('Tables', []).append(f'{title} [{", ".join(col_labels)}]')
                    elif wt == 'statusbar':
                        tag_labels = [t.get('label', '') for t in (cfg.get('tags') or []) if isinstance(t, dict)]
                        widget_types.setdefault('Status', []).append(f'{", ".join(tag_labels)}')
                    elif wt == 'text':
                        if title:
                            widget_types.setdefault('Headings', []).append(title[:60])
                lines.append(f'  Tab "{tab_label}":')
                for cat, items in widget_types.items():
                    lines.append(f'    {cat}: {", ".join(items[:8])}')
            parts.append('\n'.join(lines))
        elif lc.get('widgets'):
            widgets = lc['widgets']
            lines = [f'Report "{name}" (Dashboard): {len(widgets)} widgets']
            parts.append('\n'.join(lines))

    return '\n\n'.join(parts) if parts else ''


def _generate_ai_summary(report_names, tag_data, from_dt, to_dt, layout_configs=None):
    """
    Generate AI summary for distribution email.
    Args:
        report_names: list of report name strings
        tag_data: dict {tag_name_or_namespaced: value} — combined from all reports
        from_dt, to_dt: datetime range
        layout_configs: dict {report_name: layout_config} for report structure context
    Returns: summary text string or None
    """
    global _ai_call_count, _ai_call_date

    # Rate limit check
    today = datetime.now().date()
    if _ai_call_date != today:
        _ai_call_count = 0
        _ai_call_date = today
    if _ai_call_count >= _AI_DAILY_CAP:
        logger.warning("AI daily rate limit reached (%d calls)", _AI_DAILY_CAP)
        return None

    # Load all config from DB
    get_conn = _get_db_connection()
    ai_config = {}

    try:
        with closing(get_conn()) as conn:
            actual = conn._conn if hasattr(conn, '_conn') else conn
            cur = actual.cursor(cursor_factory=RealDictCursor)

            cur.execute("SELECT key, value FROM hercules_ai_config")
            for row in cur.fetchall():
                val = row['value'] if isinstance(row['value'], dict) else json.loads(row['value'])
                ai_config[row['key']] = val.get('value', val)

            provider = ai_config.get('ai_provider', 'cloud')
            if provider == 'cloud' and not ai_config.get('llm_api_key'):
                logger.warning("No LLM API key configured, skipping AI summary")
                return None

            # Load tracked profiles for tags in tag_data
            # Strip namespace prefixes for lookup
            raw_tags = set()
            for k in tag_data:
                if '::' in k:
                    raw_tags.add(k.split('::', 1)[1])
                else:
                    raw_tags.add(k)

            tag_list = list(raw_tags)
            if not tag_list:
                return None

            cur.execute("""
                SELECT tag_name, label, tag_type, line_name
                FROM hercules_ai_tag_profiles
                WHERE tag_name = ANY(%s) AND is_tracked = true
            """, (tag_list,))
            profile_map = {r['tag_name']: r for r in cur.fetchall()}
            actual.commit()

    except Exception as e:
        logger.warning("Failed to load AI config/profiles: %s", e)
        return None

    # Tag significance filter (max 30 tags to LLM)
    data_rows = []
    for key, value in tag_data.items():
        agg_prefix = ''
        if '::' in key:
            agg_prefix, tag_name = key.split('::', 1)
        else:
            tag_name = key
            agg_prefix = 'last'
        prof = profile_map.get(tag_name, {})
        if not prof:
            continue
        data_rows.append({
            'label': prof.get('label') or tag_name,
            'tag_type': prof.get('tag_type', 'unknown'),
            'value': value,
            'aggregation': agg_prefix,
            'line': prof.get('line_name', ''),
            'tag_name': tag_name,
        })

    if not data_rows:
        return None

    if len(data_rows) > 30:
        # Prioritize: counters first, then rate by abs value, booleans at 0, rest by abs delta
        counters = [r for r in data_rows if r['tag_type'] == 'counter']
        rates = sorted(
            [r for r in data_rows if r['tag_type'] == 'rate'],
            key=lambda r: abs(float(r['value'])) if _is_number(r['value']) else 0,
            reverse=True
        )
        booleans_zero = [r for r in data_rows if r['tag_type'] == 'boolean' and _is_zero(r['value'])]
        rest = [r for r in data_rows if r not in counters and r not in rates and r not in booleans_zero]
        if rest:
            values = [float(r['value']) for r in rest if _is_number(r['value'])]
            if values:
                mean_val = sum(values) / len(values)
                rest.sort(key=lambda r: abs(float(r['value']) - mean_val) if _is_number(r['value']) else 0, reverse=True)

        data_rows = (counters + rates + booleans_zero + rest)[:30]

    # Build structured table with aggregation context
    data_lines = []
    for r in data_rows:
        data_lines.append(f"{r['label']} | {r['tag_type']} | {r['value']} | {r['aggregation']} | {r['line']}")
    structured_data = '\n'.join(data_lines)

    names_str = ', '.join(report_names) if isinstance(report_names, list) else str(report_names)
    time_from = from_dt.strftime('%Y-%m-%d %H:%M')
    time_to = to_dt.strftime('%Y-%m-%d %H:%M')

    # Build report structure context from layout_configs
    report_context = _extract_report_context(layout_configs) if layout_configs else ''

    prompt = f"""You analyze industrial production and energy data for mill/plant managers. Be direct, specific, and useful.

REPORTS: {names_str}
PERIOD: {time_from} to {time_to}
"""
    if report_context:
        prompt += f"""
REPORT STRUCTURE (what the report sections and columns mean):
{report_context}
"""

    prompt += f"""
TAG DATA (Label | Type | Value | Aggregation | Production Line):
{structured_data}

AGGREGATION KEY:
- delta = amount produced/consumed during the period (this IS the production figure)
- first = meter reading at start of period
- last = meter reading at end of period (or current value)
- avg/sum/min/max = statistical aggregation over the period

Write a smart summary using this format:

**{names_str}** — {{one-line verdict: running normally / reduced output / line stopped / no data}}

• **Production**: {{cite delta values as production amounts with units — e.g. "Wheat Scale produced 125,294 kg"}}
• **Energy**: {{power consumption, energy totals, power factor — skip if no energy data}}
• **Status**: {{equipment on/off, only if notable — skip if all normal}}
• **Alerts**: {{zero production, zero flow rates, abnormal values — or "None"}}

Rules:
- Delta values ARE the production amounts — present them as "X produced Y kg" not as raw meter readings.
- First/last values are meter start/end readings — do NOT cite these as production amounts.
- Use the Label column (not raw tag names) when referring to tags.
- Maximum 4 bullet points. Each bullet under 25 words.
- Format numbers with thousand separators (e.g. 1,234,567 kg not 1234567.0 kg).
- Round decimals: 0 for totalizers/energy, 1 for rates/percentages.
- Only cite numbers from the data. Never invent or estimate.
- N/A or missing values = "no data" — do not guess why.
- Skip any bullet with nothing to report.
- No paragraphs. No filler. No recommendations. No greetings."""

    try:
        import ai_provider
        result = ai_provider.generate(prompt, ai_config)
        if result:
            _ai_call_count += 1
        return result
    except Exception as e:
        logger.warning("AI summary API call failed: %s", e)
        return None


def _is_number(val):
    try:
        float(val)
        return True
    except (TypeError, ValueError):
        return False


def _is_zero(val):
    try:
        return float(val) == 0
    except (TypeError, ValueError):
        return False


def _format_summary_html(summary):
    """Convert AI summary markdown into a visually rich email-safe HTML card.

    Parses the structured AI output:
      Line 1: **Report** — verdict
      Bullets: • **Label**: content
    """
    import re
    escaped = html_escape(summary)
    lines = [l.strip() for l in escaped.split('\n') if l.strip()]
    if not lines:
        return ''

    # ── Parse verdict line (first line: **Name** — verdict) ──
    verdict_html = ''
    first = lines[0]
    vm = re.match(r'\*\*(.+?)\*\*\s*[—–-]\s*(.+)', first)
    if vm:
        report_name = vm.group(1)
        verdict_text = vm.group(2).strip()
        # Determine status color from verdict keywords
        vl = verdict_text.lower()
        if any(w in vl for w in ('stopped', 'no data', 'offline', 'down')):
            dot_color, dot_bg = '#dc2626', '#fef2f2'
        elif any(w in vl for w in ('reduced', 'low', 'partial', 'warning')):
            dot_color, dot_bg = '#d97706', '#fffbeb'
        else:
            dot_color, dot_bg = '#059669', '#ecfdf5'
        verdict_html = (
            f'<table width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:14px"><tr>'
            f'<td style="vertical-align:middle;width:10px;padding-right:10px">'
            f'<div style="width:10px;height:10px;border-radius:50%;background:{dot_color}"></div></td>'
            f'<td style="vertical-align:middle">'
            f'<span style="font-size:15px;font-weight:700;color:#0f172a">{report_name}</span>'
            f'<span style="font-size:13px;color:#64748b;margin-left:8px">— {verdict_text}</span>'
            f'</td></tr></table>'
        )
        lines = lines[1:]

    # ── Parse bullet lines ──
    bullet_rows = []
    # Map bullet labels to icons/colors
    BULLET_STYLE = {
        'production': {'icon': '📦', 'accent': '#0369a1', 'bg': '#f0f9ff'},
        'energy':     {'icon': '⚡', 'accent': '#7c3aed', 'bg': '#f5f3ff'},
        'status':     {'icon': '⚙', 'accent': '#0d9488', 'bg': '#f0fdfa'},
        'alerts':     {'icon': '⚠', 'accent': '#dc2626', 'bg': '#fef2f2'},
        'flow':       {'icon': '💧', 'accent': '#0284c7', 'bg': '#f0f9ff'},
    }

    for line in lines:
        # Match: • **Label**: content  OR  • **Label** content
        bm = re.match(r'[•\-]\s*\*\*(.+?)\*\*:?\s*(.*)', line)
        if bm:
            label = bm.group(1).strip()
            content = bm.group(2).strip()
            # Bold any remaining **text** in content
            content = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', content)

            # Pick style by label
            label_key = label.lower().split()[0] if label else ''
            style = BULLET_STYLE.get(label_key, {'icon': '•', 'accent': '#475569', 'bg': '#f8fafc'})

            # Handle "None" alerts specially
            if label_key == 'alerts' and content.lower().strip() in ('none', 'none.', '—'):
                style = {**style, 'accent': '#059669', 'bg': '#ecfdf5', 'icon': '✓'}
                content = '<span style="color:#059669;font-weight:600">None</span>'

            bullet_rows.append(
                f'<tr><td style="padding:8px 12px;background:{style["bg"]};border-radius:6px;margin-bottom:4px">'
                f'<table cellpadding="0" cellspacing="0" width="100%"><tr>'
                f'<td style="vertical-align:top;width:24px;font-size:14px;padding-top:1px">{style["icon"]}</td>'
                f'<td style="vertical-align:top">'
                f'<div style="font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:0.04em;color:{style["accent"]};margin-bottom:2px">{label}</div>'
                f'<div style="font-size:13px;color:#1e293b;line-height:1.5">{content}</div>'
                f'</td></tr></table>'
                f'</td></tr>'
                f'<tr><td style="height:4px"></td></tr>'
            )
        elif line:
            # Non-bullet line (rare) — render as plain text
            rendered = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', line)
            bullet_rows.append(f'<tr><td style="padding:4px 0;font-size:13px;color:#1e293b">{rendered}</td></tr>')

    bullets_html = f'<table width="100%" cellpadding="0" cellspacing="0">{"".join(bullet_rows)}</table>' if bullet_rows else ''

    return verdict_html + bullets_html


def _prepend_summary_to_email(summary, email_html):
    """Insert styled AI summary card into the email HTML."""
    formatted = _format_summary_html(summary)
    if not formatted:
        return email_html

    summary_row = (
        '<tr><td style="padding:0 32px 24px 32px">'
        # Outer card
        '<div style="background:#ffffff;border:1px solid #e2e8f0;border-radius:10px;'
        'padding:20px 24px;box-shadow:0 1px 3px rgba(0,0,0,0.06)">'
        # Header bar
        '<table width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:14px;'
        'border-bottom:1px solid #e2e8f0;padding-bottom:10px"><tr>'
        '<td style="vertical-align:middle">'
        '<span style="font-size:12px;font-weight:800;text-transform:uppercase;'
        'letter-spacing:0.08em;color:#0369a1">✦ AI Insights</span></td>'
        '<td style="text-align:right;vertical-align:middle">'
        '<span style="font-size:10px;font-weight:600;color:#94a3b8;'
        'background:#f8fafc;padding:3px 10px;border-radius:99px">Hercules AI</span></td>'
        '</tr></table>'
        # Content
        f'{formatted}'
        '</div>'
        '</td></tr>'
    )

    # Insert before the <!-- Footer --> comment or the footer <tr>
    marker = '<!-- Footer -->'
    idx = email_html.find(marker)
    if idx >= 0:
        return email_html[:idx] + summary_row + '\n\n  ' + email_html[idx:]
    idx = email_html.lower().find('</table>\n</td></tr>')
    if idx >= 0:
        return email_html[:idx] + summary_row + '\n' + email_html[idx:]
    idx = email_html.lower().find('</body>')
    if idx >= 0:
        return email_html[:idx] + summary_row + email_html[idx:]
    return email_html + summary_row
