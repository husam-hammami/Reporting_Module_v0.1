"""
Tags Blueprint

API endpoints for managing tags (PLC tags, formulas, mappings, manual inputs).
"""

import logging
import json
import csv
import io
import re
import time
from flask import Blueprint, jsonify, request
from contextlib import closing
import psycopg2
from psycopg2.extras import RealDictCursor
from utils.plc_parser import parse_plc_address
from utils.tag_reader import read_tag_value, read_all_tags
from plc_utils import connect_to_plc_fast

logger = logging.getLogger(__name__)


def _sanitize_tag_name(name):
    """Sanitize tag name: replace spaces with underscores, strip special chars."""
    clean = name.replace(' ', '_')
    clean = re.sub(r'[^a-zA-Z0-9_\-.]', '', clean)
    return clean or 'unnamed_tag'


tags_bp = Blueprint('tags_bp', __name__)


def _get_db_connection():
    """Helper function to get database connection, avoiding circular imports"""
    import sys
    if 'app' in sys.modules:
        app_module = sys.modules['app']
        get_db_connection = getattr(app_module, 'get_db_connection', None)
        if get_db_connection is None:
            raise ImportError("get_db_connection not found in app module")
        return get_db_connection
    else:
        from app import get_db_connection
        return get_db_connection


@tags_bp.route('/tags', methods=['GET'])
def get_all_tags():
    """Get all active tags (with optional filter)"""
    try:
        get_db_connection_func = _get_db_connection()
        source_type = request.args.get('source_type')
        is_active = request.args.get('is_active', 'true').lower() == 'true'
        
        # Use closing() context manager to ensure proper cleanup
        with closing(get_db_connection_func()) as conn:
            # Handle both PooledConnection and regular connection
            actual_conn = conn._conn if hasattr(conn, '_conn') else conn
            
            cursor = actual_conn.cursor(cursor_factory=RealDictCursor)
            
            query = """
                SELECT id, tag_name, display_name, source_type, 
                       db_number, "offset", data_type, bit_position,
                       string_length, byte_swap, unit, scaling, 
                       decimal_places, formula, mapping_name, value_formula,
                       description, is_active, is_bin_tag, 
                       activation_tag_name, activation_condition, activation_value,
                       created_at, updated_at
                FROM tags 
                WHERE 1=1
            """
            params = []
            
            if is_active is not None:
                query += " AND is_active = %s"
                params.append(is_active)
            
            if source_type:
                query += " AND source_type = %s"
                params.append(source_type)
            
            query += " ORDER BY tag_name"
            
            cursor.execute(query, params)
            tags = cursor.fetchall()
            
            # Convert to list of dicts
            result = []
            for tag in tags:
                tag_dict = dict(tag)
                # Format PLC address for display
                if tag_dict.get('db_number') is not None and tag_dict.get('offset') is not None:
                    if tag_dict.get('bit_position') is not None:
                        tag_dict['plc_address'] = f"DB{tag_dict['db_number']}.{tag_dict['offset']}.{tag_dict['bit_position']}"
                    else:
                        tag_dict['plc_address'] = f"DB{tag_dict['db_number']}.{tag_dict['offset']}"
                else:
                    tag_dict['plc_address'] = None
                
                result.append(tag_dict)
            
            logger.debug(f"Returned {len(result)} tags")
            return jsonify({
                'status': 'success',
                'tags': result
            })
    
    except psycopg2.OperationalError as e:
        logger.error(f"Database operational error getting tags: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': f'Database connection error: {str(e)}'}), 500
    except psycopg2.Error as e:
        logger.error(f"Database error getting tags: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': f'Database error: {str(e)}'}), 500
    except Exception as e:
        logger.error(f"Error getting tags: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': str(e)}), 500


@tags_bp.route('/tags', methods=['POST'])
def create_tag():
    """Create a new tag"""
    try:
        logger.info("=== CREATE TAG REQUEST START ===")
        logger.info(f"Request data: {request.get_json()}")
        
        data = request.get_json()
        
        if not data:
            logger.error("No data received in request")
            return jsonify({'status': 'error', 'message': 'No data received'}), 400
        
        if not data.get('tag_name'):
            logger.error("tag_name is missing")
            return jsonify({'status': 'error', 'message': 'tag_name is required'}), 400
        
        tag_name = data['tag_name'].strip()
        source_type = data.get('source_type', 'PLC')
        
        # Parse PLC address if provided
        db_number = None
        offset = None
        bit_position = None
        
        if source_type == 'PLC':
            plc_address = data.get('plc_address')
            if not plc_address:
                return jsonify({'status': 'error', 'message': 'plc_address is required for PLC tags'}), 400
            
            try:
                parsed = parse_plc_address(plc_address)
                db_number = parsed['db_number']
                offset = parsed['offset']
                bit_position = parsed.get('bit')
            except ValueError as e:
                return jsonify({'status': 'error', 'message': str(e)}), 400
        
        # Validate data type
        data_type = data.get('data_type', 'REAL')
        if data_type not in ['BOOL', 'INT', 'DINT', 'REAL', 'STRING']:
            return jsonify({'status': 'error', 'message': f'Invalid data_type: {data_type}'}), 400
        
        # Validate bit_position for BOOL
        if data_type == 'BOOL' and bit_position is None:
            bit_pos = data.get('bit_position')
            if bit_pos is not None and bit_pos != '':
                try:
                    bit_position = int(bit_pos)
                except (ValueError, TypeError):
                    bit_position = 0
            else:
                bit_position = 0
        elif data_type != 'BOOL':
            # For non-BOOL types, bit_position should be None
            bit_position = None
        
        logger.info(f"Attempting to connect to database...")
        try:
            get_db_connection_func = _get_db_connection()
            if get_db_connection_func is None:
                raise ImportError("get_db_connection not available")
            conn = get_db_connection_func()
            logger.info("Database connection obtained")
        except Exception as conn_error:
            logger.error(f"Failed to get database connection: {conn_error}", exc_info=True)
            return jsonify({
                'status': 'error',
                'message': f'Database connection failed: {str(conn_error)}'
            }), 500
        
        try:
            with closing(conn) as conn:
                # Use regular cursor for INSERT operations (not RealDictCursor)
                cursor = conn.cursor(cursor_factory=None)
                logger.info("Database cursor created")
                
                # Check if tag_name already exists
                logger.info(f"Checking if tag '{tag_name}' already exists...")
                cursor.execute("SELECT id FROM tags WHERE tag_name = %s", (tag_name,))
                existing = cursor.fetchone()
                if existing:
                    logger.warning(f"Tag '{tag_name}' already exists")
                    return jsonify({'status': 'error', 'message': f'Tag name "{tag_name}" already exists'}), 400
                logger.info(f"Tag '{tag_name}' is available")
                
                # Insert tag
                # Prepare values with proper type handling
                string_length_val = None
                if data_type == 'STRING':
                    str_len = data.get('string_length', 40)
                    string_length_val = int(str_len) if str_len and str_len != '' else 40
                
                byte_swap_val = None
                if data_type == 'REAL':
                    byte_swap_val = bool(data.get('byte_swap', False))  # Default to False (big-endian)
                
                is_active_val = bool(data.get('is_active', True))
                
                # Ensure bit_position is None for non-BOOL types
                if data_type != 'BOOL':
                    bit_position = None
                
                # Log the values being inserted for debugging
                logger.debug(f"Inserting tag with values: tag_name={tag_name}, db_number={db_number}, offset={offset}, data_type={data_type}, bit_position={bit_position}")
                
                try:
                    # Get bin activation fields
                    is_bin_tag = bool(data.get('is_bin_tag', False))
                    activation_tag_name = data.get('activation_tag_name') or None
                    activation_condition = data.get('activation_condition') or None
                    activation_value = data.get('activation_value') or None
                    
                    cursor.execute("""
                        INSERT INTO tags (
                            tag_name, display_name, source_type, db_number, "offset",
                            data_type, bit_position, string_length, byte_swap,
                            unit, scaling, decimal_places, formula, mapping_name,
                            description, is_active, is_bin_tag,
                            activation_tag_name, activation_condition, activation_value,
                            value_formula
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        RETURNING id
                    """, (
                        tag_name,
                        data.get('display_name', tag_name),
                        source_type,
                        db_number,
                        offset,
                        data_type,
                        bit_position,
                        string_length_val,
                        byte_swap_val,
                        data.get('unit', ''),
                        float(data.get('scaling', 1.0)),
                        int(data.get('decimal_places', 2)),
                        data.get('formula') or None,
                        data.get('mapping_name') or None,
                        data.get('description', ''),
                        is_active_val,
                        is_bin_tag,
                        activation_tag_name,
                        activation_condition,
                        activation_value,
                        data.get('value_formula') or None
                    ))
                    
                    result = cursor.fetchone()
                    if not result:
                        raise Exception("INSERT did not return an ID")
                    # Handle both dict (RealDictCursor) and tuple results
                    if isinstance(result, dict):
                        tag_id = result.get('id') or result.get('ID')
                    else:
                        tag_id = result[0]
                    if tag_id is None:
                        raise Exception("INSERT did not return a valid ID")
                    
                    conn.commit()
                    logger.info(f"Created tag: {tag_name} (ID: {tag_id})")

                    # Auto-register in emulator if demo mode is active
                    try:
                        from demo_mode import get_demo_mode
                        if get_demo_mode():
                            from plc_data_source import register_tag_in_emulator
                            register_tag_in_emulator(tag_name, source_type, db_number, offset, data_type, data.get('unit', ''))
                    except Exception as emu_err:
                        logger.debug("Emulator auto-register skipped: %s", emu_err)

                    return jsonify({
                        'status': 'success',
                        'tag_id': tag_id,
                        'message': f'Tag "{tag_name}" created successfully'
                    }), 201
                except Exception as db_error:
                    conn.rollback()
                    logger.error(f"Database insert error: {db_error}", exc_info=True)
                    raise
        except Exception as db_error:
            logger.error(f"Database operation error: {db_error}", exc_info=True)
            raise
    
    except psycopg2.IntegrityError as e:
        logger.error(f"Database integrity error creating tag: {e}", exc_info=True)
        error_msg = str(e) or repr(e) or 'Unknown integrity error'
        if 'unique constraint' in error_msg.lower() or 'duplicate key' in error_msg.lower():
            return jsonify({
                'status': 'error',
                'message': f'Tag name already exists'
            }), 400
        return jsonify({
            'status': 'error',
            'message': f'Database constraint violation: {error_msg}'
        }), 400
    except psycopg2.Error as e:
        logger.error(f"Database error creating tag: {e}", exc_info=True)
        error_msg = str(e) or repr(e) or 'Unknown database error'
        return jsonify({
            'status': 'error',
            'message': f'Database error: {error_msg}'
        }), 500
    except Exception as e:
        logger.error(f"Error creating tag: {e}", exc_info=True)
        import traceback
        error_details = traceback.format_exc()
        logger.error(f"Full traceback: {error_details}")
        error_msg = str(e) or repr(e) or f'Unknown error: {type(e).__name__}'
        return jsonify({
            'status': 'error', 
            'message': error_msg,
            'error_type': type(e).__name__
        }), 500


@tags_bp.route('/tags/get-values', methods=['POST'])
def get_tag_values():
    """Get current values for multiple tags from PLC, with activation checking for bin tags"""
    try:
        logger.info("=== GET TAG VALUES REQUEST START ===")
        logger.info(f"Request method: {request.method}")
        logger.info(f"Request path: {request.path}")
        logger.info(f"Request JSON: {request.get_json()}")
        
        data = request.get_json()
        if not data:
            logger.warning("No JSON data in request")
            return jsonify({
                'status': 'error',
                'message': 'No data received'
            }), 400
        
        tag_names = data.get('tag_names', [])
        logger.info(f"Requested tag names: {tag_names}")
        
        if not tag_names:
            logger.warning("Empty tag_names array")
            return jsonify({
                'status': 'error',
                'message': 'tag_names array is required'
            }), 400
        
        get_db_connection_func = _get_db_connection()
        
        try:
            # ✅ NEW: Load tag configurations to check for bin activation conditions
            with closing(get_db_connection_func()) as conn:
                cursor = conn.cursor(cursor_factory=RealDictCursor)
                cursor.execute("""
                    SELECT tag_name, is_bin_tag, activation_tag_name, 
                           activation_condition, activation_value
                    FROM tags
                    WHERE tag_name = ANY(%s)
                    AND is_active = true
                """, (tag_names,))
                tag_configs = {row['tag_name']: row for row in cursor.fetchall()}
            
            # Read all tag values from PLC
            logger.info(f"Reading values for {len(tag_names)} tags from PLC...")
            tag_values = read_all_tags(tag_names=tag_names, db_connection_func=get_db_connection_func)
            logger.info(f"Successfully read {len(tag_values)} tag values: {list(tag_values.keys())}")
            
            # ✅ NEW: Check activation conditions for bin tags
            active_tag_values = {}
            activation_tags_needed = set()
            
            # First pass: identify which activation tags we need
            for tag_name in tag_names:
                tag_config = tag_configs.get(tag_name, {})
                if tag_config.get('is_bin_tag') and tag_config.get('activation_tag_name'):
                    activation_tags_needed.add(tag_config['activation_tag_name'])
            
            # Fetch activation tag values if needed
            activation_values = {}
            if activation_tags_needed:
                logger.info(f"Fetching activation tag values: {list(activation_tags_needed)}")
                activation_values = read_all_tags(
                    tag_names=list(activation_tags_needed),
                    db_connection_func=get_db_connection_func
                )
                logger.info(f"Activation tag values: {activation_values}")
            
            # Second pass: filter bin tags based on activation conditions
            for tag_name, value in tag_values.items():
                tag_config = tag_configs.get(tag_name, {})
                
                # ✅ If this is a bin tag with activation condition, check it
                if tag_config.get('is_bin_tag') and tag_config.get('activation_tag_name'):
                    activation_tag = tag_config['activation_tag_name']
                    activation_condition = tag_config.get('activation_condition', 'equals')
                    activation_value = tag_config.get('activation_value')
                    
                    # Get activation tag value
                    activation_tag_value = activation_values.get(activation_tag)
                    
                    # ✅ Evaluate activation condition
                    is_active = evaluate_activation_condition(
                        activation_tag_value,
                        activation_condition,
                        activation_value
                    )
                    
                    # ✅ If bin is not active, set value to 0 (inactive)
                    if not is_active:
                        logger.debug(f"[Bin Activation] Tag {tag_name} (bin_id={value}) is inactive: {activation_tag}={activation_tag_value} does not meet condition {activation_condition}={activation_value}")
                        active_tag_values[tag_name] = 0
                    else:
                        active_tag_values[tag_name] = value
                        logger.debug(f"[Bin Activation] Tag {tag_name} (bin_id={value}) is active: {activation_tag}={activation_tag_value} meets condition")
                else:
                    # Not a bin tag or no activation condition, include as-is
                    active_tag_values[tag_name] = value
            
            logger.info(f"Returning {len(active_tag_values)} tag values (after activation filtering)")
            return jsonify({
                'status': 'success',
                'tag_values': active_tag_values
            })
        except Exception as e:
            logger.error(f"Error reading tag values: {e}", exc_info=True)
            return jsonify({
                'status': 'error',
                'message': f'Failed to read tag values: {str(e)}',
                'tag_values': {}
            }), 500
    
    except Exception as e:
        logger.error(f"Error in get_tag_values: {e}", exc_info=True)
        return jsonify({
            'status': 'error',
            'message': str(e),
            'tag_values': {}
        }), 500


def evaluate_activation_condition(actual_value, condition, expected_value):
    """Evaluate if activation condition is met"""
    if actual_value is None:
        return False
    
    # Convert to strings for comparison
    if isinstance(actual_value, bool):
        actual_str = str(actual_value).lower()
    else:
        actual_str = str(actual_value).lower()
    
    expected_str = str(expected_value).lower() if expected_value else ''
    
    if condition == "equals":
        return actual_str == expected_str
    elif condition == "not_equals":
        return actual_str != expected_str
    elif condition == "true":
        return bool(actual_value) is True
    elif condition == "false":
        return bool(actual_value) is False
    elif condition == "greater_than":
        try:
            return float(actual_value) > float(expected_value)
        except:
            return False
    elif condition == "less_than":
        try:
            return float(actual_value) < float(expected_value)
        except:
            return False
    else:
        return True  # Default: always active if condition unknown


@tags_bp.route('/tags/<tag_name>', methods=['GET'])
def get_tag(tag_name):
    """Get single tag details"""
    try:
        get_db_connection_func = _get_db_connection()
        with closing(get_db_connection_func()) as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            cursor.execute("""
                SELECT id, tag_name, display_name, source_type, 
                       db_number, "offset", data_type, bit_position,
                       string_length, byte_swap, unit, scaling, 
                       decimal_places, formula, mapping_name, value_formula,
                       description, is_active, is_bin_tag,
                       activation_tag_name, activation_condition, activation_value,
                       created_at, updated_at
                FROM tags 
                WHERE tag_name = %s
            """, (tag_name,))
            
            tag = cursor.fetchone()
            
            if not tag:
                return jsonify({'status': 'error', 'message': 'Tag not found'}), 404
            
            tag_dict = dict(tag)
            # Format PLC address
            if tag_dict.get('db_number') is not None and tag_dict.get('offset') is not None:
                if tag_dict.get('bit_position') is not None:
                    tag_dict['plc_address'] = f"DB{tag_dict['db_number']}.{tag_dict['offset']}.{tag_dict['bit_position']}"
                else:
                    tag_dict['plc_address'] = f"DB{tag_dict['db_number']}.{tag_dict['offset']}"
            else:
                tag_dict['plc_address'] = None
            
            return jsonify({
                'status': 'success',
                'tag': tag_dict
            })
    
    except Exception as e:
        logger.error(f"Error getting tag: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': str(e)}), 500


@tags_bp.route('/tags/<tag_name>', methods=['PUT'])
def update_tag(tag_name):
    """Update tag"""
    try:
        data = request.get_json()
        
        get_db_connection_func = _get_db_connection()
        with closing(get_db_connection_func()) as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            # Check if tag exists
            cursor.execute("SELECT * FROM tags WHERE tag_name = %s", (tag_name,))
            existing_tag = cursor.fetchone()
            
            if not existing_tag:
                return jsonify({'status': 'error', 'message': 'Tag not found'}), 404
            
            existing_dict = dict(existing_tag)
            source_type = data.get('source_type', existing_dict.get('source_type', 'PLC'))
            data_type = data.get('data_type', existing_dict.get('data_type'))
            
            # Parse PLC address if provided
            db_number = existing_dict.get('db_number')
            offset = existing_dict.get('offset')
            bit_position = existing_dict.get('bit_position')
            
            if source_type == 'PLC' and data.get('plc_address'):
                try:
                    parsed = parse_plc_address(data['plc_address'])
                    db_number = parsed['db_number']
                    offset = parsed['offset']
                    bit_position = parsed.get('bit')
                except ValueError as e:
                    return jsonify({'status': 'error', 'message': str(e)}), 400
            
            # Handle bit_position from form data for BOOL types
            if data_type == 'BOOL':
                bit_pos = data.get('bit_position')
                if bit_pos is not None and bit_pos != '':
                    try:
                        bit_position = int(bit_pos)
                        if bit_position < 0 or bit_position > 7:
                            return jsonify({'status': 'error', 'message': 'Bit Position must be 0-7 for BOOL type'}), 400
                    except (ValueError, TypeError):
                        # If bit_position is invalid, keep existing value
                        pass
                elif bit_pos == '' or bit_pos is None:
                    # If empty string or None, use 0 as default for BOOL
                    bit_position = 0
            else:
                # For non-BOOL types, bit_position should be None
                bit_position = None
            
            # Get bin activation fields
            is_bin_tag = data.get('is_bin_tag')
            if is_bin_tag is None:
                is_bin_tag = existing_dict.get('is_bin_tag', False)
            else:
                is_bin_tag = bool(is_bin_tag)
            
            activation_tag_name = data.get('activation_tag_name')
            if activation_tag_name is None:
                activation_tag_name = existing_dict.get('activation_tag_name')
            else:
                activation_tag_name = activation_tag_name if activation_tag_name else None
            
            activation_condition = data.get('activation_condition')
            if activation_condition is None:
                activation_condition = existing_dict.get('activation_condition')
            else:
                activation_condition = activation_condition if activation_condition else None
            
            activation_value = data.get('activation_value')
            if activation_value is None:
                activation_value = existing_dict.get('activation_value')
            else:
                activation_value = activation_value if activation_value else None
            
            # Update tag
            cursor.execute("""
                UPDATE tags SET
                    display_name = %s,
                    source_type = %s,
                    db_number = %s,
                    "offset" = %s,
                    data_type = %s,
                    bit_position = %s,
                    string_length = %s,
                    byte_swap = %s,
                    unit = %s,
                    scaling = %s,
                    decimal_places = %s,
                    formula = %s,
                    mapping_name = %s,
                    description = %s,
                    is_active = %s,
                    is_bin_tag = %s,
                    activation_tag_name = %s,
                    activation_condition = %s,
                    activation_value = %s
                WHERE tag_name = %s
                RETURNING id
            """, (
                data.get('display_name', existing_dict.get('display_name')),
                source_type,
                db_number,
                offset,
                data_type,
                bit_position,
                data.get('string_length', existing_dict.get('string_length')),
                data.get('byte_swap', existing_dict.get('byte_swap')),
                data.get('unit', existing_dict.get('unit', '')),
                float(data.get('scaling', existing_dict.get('scaling', 1.0))),
                int(data.get('decimal_places', existing_dict.get('decimal_places', 2))),
                data.get('formula', existing_dict.get('formula')),
                data.get('mapping_name', existing_dict.get('mapping_name')),
                data.get('description', existing_dict.get('description', '')),
                data.get('is_active', existing_dict.get('is_active', True)),
                is_bin_tag,
                activation_tag_name,
                activation_condition,
                activation_value,
                tag_name
            ))
            
            # Safely get the result
            result = cursor.fetchone()
            if not result:
                conn.rollback()
                return jsonify({'status': 'error', 'message': 'Failed to update tag'}), 500
            
            # Handle both dict (RealDictCursor) and tuple results
            if isinstance(result, dict):
                tag_id = result.get('id') or result.get('ID')
            else:
                tag_id = result[0]
            
            if tag_id is None:
                conn.rollback()
                return jsonify({'status': 'error', 'message': 'Failed to update tag'}), 500
            
            conn.commit()

            logger.info(f"Updated tag: {tag_name} (ID: {tag_id})")

            # Auto-register in emulator if demo mode is active
            try:
                from demo_mode import get_demo_mode
                if get_demo_mode():
                    from plc_data_source import register_tag_in_emulator
                    register_tag_in_emulator(tag_name, source_type, db_number, offset, data_type, data.get('unit', ''))
            except Exception as emu_err:
                logger.debug("Emulator auto-register on update skipped: %s", emu_err)

            return jsonify({
                'status': 'success',
                'message': f'Tag "{tag_name}" updated successfully'
            })
    
    except Exception as e:
        logger.error(f"Error updating tag: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': str(e)}), 500


@tags_bp.route('/tags/<tag_name>', methods=['DELETE'])
def delete_tag(tag_name):
    """Delete tag (soft delete: set is_active=false)"""
    try:
        get_db_connection_func = _get_db_connection()
        with closing(get_db_connection_func()) as conn:
            cursor = conn.cursor()
            
            # Check if tag exists
            cursor.execute("SELECT id FROM tags WHERE tag_name = %s", (tag_name,))
            if not cursor.fetchone():
                return jsonify({'status': 'error', 'message': 'Tag not found'}), 404
            
            # Soft delete
            cursor.execute("UPDATE tags SET is_active = false WHERE tag_name = %s", (tag_name,))
            conn.commit()
            
            logger.info(f"Deleted tag: {tag_name}")
            
            return jsonify({
                'status': 'success',
                'message': f'Tag "{tag_name}" deleted successfully'
            })
    
    except Exception as e:
        logger.error(f"Error deleting tag: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': str(e)}), 500


@tags_bp.route('/tags/<tag_name>/test', methods=['GET'])
def test_tag(tag_name):
    """Test reading a tag from PLC"""
    try:
        get_db_connection_func = _get_db_connection()
        with closing(get_db_connection_func()) as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            cursor.execute("SELECT * FROM tags WHERE tag_name = %s", (tag_name,))
            tag = cursor.fetchone()
            
            if not tag:
                return jsonify({'status': 'error', 'message': 'Tag not found'}), 404
            
            tag_dict = dict(tag)
            
            if tag_dict['source_type'] != 'PLC':
                return jsonify({'status': 'error', 'message': 'Tag is not a PLC tag'}), 400
            
            if not tag_dict.get('db_number') or tag_dict.get('offset') is None:
                return jsonify({'status': 'error', 'message': 'Tag missing PLC address configuration'}), 400
            
            # Read from PLC
            try:
                plc = connect_to_plc_fast()
                value = read_tag_value(plc, tag_dict)
                
                if value is None:
                    return jsonify({
                        'status': 'error',
                        'message': 'Failed to read tag from PLC. Check PLC connection and address.'
                    }), 500
                
                # ✅ NEW: Apply value formula if provided, otherwise use scaling (backward compatibility)
                value_formula = tag_dict.get('value_formula')
                if value_formula and value_formula.strip():
                    from utils.tag_reader import evaluate_value_formula
                    final_value = evaluate_value_formula(value_formula, value)
                else:
                    scaling = float(tag_dict.get('scaling', 1.0))
                    final_value = value * scaling
                
                return jsonify({
                    'status': 'success',
                    'tag_name': tag_name,
                    'value': final_value,
                    'raw_value': value,
                    'unit': tag_dict.get('unit', ''),
                    'plc_address': f"DB{tag_dict['db_number']}.{tag_dict['offset']}"
                })
            
            except Exception as e:
                logger.error(f"PLC read error for tag {tag_name}: {e}", exc_info=True)
                return jsonify({
                    'status': 'error',
                    'message': f'PLC read failed: {str(e)}'
                }), 500
    
    except Exception as e:
        logger.error(f"Error testing tag: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': str(e)}), 500


@tags_bp.route('/tags/bulk-import', methods=['POST'])
def bulk_import_tags():
    """Import tags from JSON"""
    try:
        data = request.get_json()
        tags = data.get('tags', [])
        
        if not tags:
            return jsonify({'status': 'error', 'message': 'No tags provided'}), 400
        
        imported = 0
        errors = []
        
        get_db_connection_func = _get_db_connection()
        with closing(get_db_connection_func()) as conn:
            cursor = conn.cursor()
            
            for tag_data in tags:
                try:
                    tag_name = tag_data.get('tag_name')
                    if not tag_name:
                        errors.append({'tag': tag_data, 'error': 'Missing tag_name'})
                        continue
                    
                    source_type = tag_data.get('source_type', 'PLC')
                    db_number = None
                    offset = None
                    bit_position = None
                    
                    if source_type == 'PLC' and tag_data.get('plc_address'):
                        try:
                            parsed = parse_plc_address(tag_data['plc_address'])
                            db_number = parsed['db_number']
                            offset = parsed['offset']
                            bit_position = parsed.get('bit')
                        except ValueError as e:
                            errors.append({'tag': tag_name, 'error': str(e)})
                            continue
                    
                    # Check if exists
                    cursor.execute("SELECT id FROM tags WHERE tag_name = %s", (tag_name,))
                    if cursor.fetchone():
                        # Update existing
                        cursor.execute("""
                            UPDATE tags SET
                                display_name = %s, source_type = %s, db_number = %s,
                                "offset" = %s, data_type = %s, bit_position = %s,
                                string_length = %s, byte_swap = %s, unit = %s,
                                scaling = %s, decimal_places = %s, description = %s,
                                is_active = %s
                            WHERE tag_name = %s
                        """, (
                            tag_data.get('display_name', tag_name),
                            source_type, db_number, offset,
                            tag_data.get('data_type', 'REAL'),
                            bit_position,
                            tag_data.get('string_length', 40),
                            tag_data.get('byte_swap', True),
                            tag_data.get('unit', ''),
                            float(tag_data.get('scaling', 1.0)),
                            int(tag_data.get('decimal_places', 2)),
                            tag_data.get('description', ''),
                            tag_data.get('is_active', True),
                            tag_name
                        ))
                    else:
                        # Insert new
                        cursor.execute("""
                            INSERT INTO tags (
                                tag_name, display_name, source_type, db_number, "offset",
                                data_type, bit_position, string_length, byte_swap,
                                unit, scaling, decimal_places, description, is_active
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            tag_name,
                            tag_data.get('display_name', tag_name),
                            source_type, db_number, offset,
                            tag_data.get('data_type', 'REAL'),
                            bit_position,
                            tag_data.get('string_length', 40),
                            tag_data.get('byte_swap', True),
                            tag_data.get('unit', ''),
                            float(tag_data.get('scaling', 1.0)),
                            int(tag_data.get('decimal_places', 2)),
                            tag_data.get('description', ''),
                            tag_data.get('is_active', True)
                        ))
                    
                    imported += 1
                
                except Exception as e:
                    errors.append({'tag': tag_data.get('tag_name', 'Unknown'), 'error': str(e)})
            
            conn.commit()

        # Auto-register all imported tags in emulator if demo mode is active
        try:
            from demo_mode import get_demo_mode
            if get_demo_mode():
                from plc_data_source import register_tag_in_emulator
                for tag_data in tags:
                    t_name = tag_data.get('tag_name')
                    if not t_name:
                        continue
                    t_source = tag_data.get('source_type', 'PLC')
                    t_db = None
                    t_off = None
                    if t_source == 'PLC' and tag_data.get('plc_address'):
                        try:
                            p = parse_plc_address(tag_data['plc_address'])
                            t_db = p['db_number']
                            t_off = p['offset']
                        except ValueError:
                            continue
                    register_tag_in_emulator(t_name, t_source, t_db, t_off, tag_data.get('data_type', 'REAL'), tag_data.get('unit', ''))
        except Exception as emu_err:
            logger.debug("Emulator auto-register on bulk import skipped: %s", emu_err)

        return jsonify({
            'status': 'success',
            'imported': imported,
            'errors': errors,
            'message': f'Imported {imported} tags'
        })
    
    except Exception as e:
        logger.error(f"Error bulk importing tags: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': str(e)}), 500


@tags_bp.route('/tags/seed', methods=['POST'])
def seed_demo_tags():
    """Seed the database with demo PLC tags, groups, and mappings for development."""
    try:
        SEED_TAGS = [
            {'tag_name': 'Temperature_1', 'display_name': 'Temperature Sensor 1', 'source_type': 'PLC', 'plc_address': 'DB2099.0', 'data_type': 'REAL', 'unit': '°C', 'description': 'Main process temperature', 'decimal_places': 1},
            {'tag_name': 'Pressure_1', 'display_name': 'Pressure Sensor 1', 'source_type': 'PLC', 'plc_address': 'DB2099.4', 'data_type': 'REAL', 'unit': 'bar', 'description': 'System pressure', 'decimal_places': 2},
            {'tag_name': 'Flow_Rate_1', 'display_name': 'Flow Rate', 'source_type': 'PLC', 'plc_address': 'DB2099.8', 'data_type': 'REAL', 'unit': 'm³/h', 'description': 'Main flow rate', 'decimal_places': 1},
            {'tag_name': 'Motor_Speed_1', 'display_name': 'Motor Speed', 'source_type': 'PLC', 'plc_address': 'DB2099.12', 'data_type': 'REAL', 'unit': 'RPM', 'description': 'Main motor speed', 'decimal_places': 0},
            {'tag_name': 'Level_Tank_1', 'display_name': 'Tank Level', 'source_type': 'PLC', 'plc_address': 'DB2099.16', 'data_type': 'REAL', 'unit': '%', 'description': 'Storage tank level', 'decimal_places': 1},
            {'tag_name': 'Power_Consumption', 'display_name': 'Power Consumption', 'source_type': 'PLC', 'plc_address': 'DB1603.392', 'data_type': 'REAL', 'unit': 'kW', 'description': 'Total power draw', 'decimal_places': 2},
            {'tag_name': 'Vibration_1', 'display_name': 'Vibration Sensor', 'source_type': 'PLC', 'plc_address': 'DB2099.20', 'data_type': 'REAL', 'unit': 'mm/s', 'description': 'Motor vibration', 'decimal_places': 2},
            {'tag_name': 'Weight_Scale_1', 'display_name': 'Scale Weight', 'source_type': 'PLC', 'plc_address': 'DB499.0', 'data_type': 'REAL', 'unit': 'kg', 'description': 'Product weight', 'decimal_places': 1},
            {'tag_name': 'Mill_Throughput', 'display_name': 'Mill Throughput', 'source_type': 'PLC', 'plc_address': 'DB2099.24', 'data_type': 'REAL', 'unit': 't/h', 'description': 'Production throughput', 'decimal_places': 2},
            {'tag_name': 'Flour_Extraction', 'display_name': 'Flour Extraction', 'source_type': 'PLC', 'plc_address': 'DB2099.28', 'data_type': 'REAL', 'unit': '%', 'description': 'Flour extraction rate', 'decimal_places': 2},
            {'tag_name': 'Bran_Extraction', 'display_name': 'Bran Extraction', 'source_type': 'PLC', 'plc_address': 'DB2099.32', 'data_type': 'REAL', 'unit': '%', 'description': 'Bran extraction rate', 'decimal_places': 2},
            {'tag_name': 'Water_Used', 'display_name': 'Total Water Used', 'source_type': 'PLC', 'plc_address': 'DB199.564', 'data_type': 'REAL', 'unit': 'L', 'description': 'Water consumption', 'decimal_places': 1},
            # Formula tags
            {'tag_name': 'MillingLossFormula', 'display_name': 'Milling Loss', 'source_type': 'Formula', 'data_type': 'REAL', 'unit': '%', 'description': 'Calculated milling loss: 100 - Flour - Bran', 'formula': '100 - {Flour_Extraction} - {Bran_Extraction}', 'decimal_places': 2},
            {'tag_name': 'FlowRate_Avg', 'display_name': 'Avg Flow Rate', 'source_type': 'Formula', 'data_type': 'REAL', 'unit': 'm³/h', 'description': 'Averaged flow rate', 'formula': '{Flow_Rate_1}', 'decimal_places': 1},
        ]

        SEED_GROUPS = [
            {'group_name': 'Process Sensors', 'description': 'Core process measurement sensors', 'tags': ['Temperature_1', 'Pressure_1', 'Flow_Rate_1', 'Level_Tank_1']},
            {'group_name': 'Production KPIs', 'description': 'Key production indicators', 'tags': ['Mill_Throughput', 'Flour_Extraction', 'Bran_Extraction', 'MillingLossFormula']},
            {'group_name': 'Utilities', 'description': 'Power, water, and utilities', 'tags': ['Power_Consumption', 'Water_Used']},
            {'group_name': 'Mechanical', 'description': 'Motor and vibration monitoring', 'tags': ['Motor_Speed_1', 'Vibration_1', 'Weight_Scale_1']},
        ]

        get_db_connection_func = _get_db_connection()
        imported_tags = 0
        imported_groups = 0

        with closing(get_db_connection_func()) as conn:
            actual_conn = conn._conn if hasattr(conn, '_conn') else conn
            cursor = actual_conn.cursor(cursor_factory=RealDictCursor)

            # Seed tags
            for t in SEED_TAGS:
                # Parse PLC address
                db_number = None
                offset_val = None
                bit_position = None
                if t.get('source_type') == 'PLC' and t.get('plc_address'):
                    try:
                        parsed = parse_plc_address(t['plc_address'])
                        db_number = parsed['db_number']
                        offset_val = parsed['offset']
                        bit_position = parsed.get('bit')
                    except Exception:
                        pass

                cursor.execute("SELECT 1 FROM tags WHERE tag_name = %s", (t['tag_name'],))
                if cursor.fetchone():
                    continue  # skip existing

                cursor.execute("""
                    INSERT INTO tags (tag_name, display_name, source_type, db_number, "offset",
                        data_type, bit_position, unit, decimal_places, description, formula, is_active)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    t['tag_name'], t.get('display_name', t['tag_name']), t.get('source_type', 'PLC'),
                    db_number, offset_val, t.get('data_type', 'REAL'), bit_position,
                    t.get('unit', ''), int(t.get('decimal_places', 2)), t.get('description', ''),
                    t.get('formula', ''), True
                ))
                imported_tags += 1

            actual_conn.commit()

            # Seed tag groups
            for g in SEED_GROUPS:
                cursor.execute("SELECT id FROM tag_groups WHERE group_name = %s", (g['group_name'],))
                existing = cursor.fetchone()
                if existing:
                    group_id = existing['id']
                else:
                    cursor.execute("""
                        INSERT INTO tag_groups (group_name, description, is_active)
                        VALUES (%s, %s, %s) RETURNING id
                    """, (g['group_name'], g['description'], True))
                    group_id = cursor.fetchone()['id']
                    imported_groups += 1

                # Assign tags
                for tag_name in g.get('tags', []):
                    cursor.execute("SELECT id FROM tags WHERE tag_name = %s", (tag_name,))
                    tag_row = cursor.fetchone()
                    if tag_row:
                        cursor.execute("""
                            INSERT INTO tag_group_members (group_id, tag_id)
                            VALUES (%s, %s) ON CONFLICT DO NOTHING
                        """, (group_id, tag_row['id']))

            actual_conn.commit()

        return jsonify({
            'status': 'success',
            'message': f'Seeded {imported_tags} tags and {imported_groups} groups',
            'tags_created': imported_tags,
            'groups_created': imported_groups
        })

    except Exception as e:
        logger.error(f"Error seeding demo tags: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': str(e)}), 500


@tags_bp.route('/tags/export', methods=['GET'])
def export_tags():
    """Export all tags to JSON"""
    try:
        get_db_connection_func = _get_db_connection()
        with closing(get_db_connection_func()) as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            cursor.execute("""
                SELECT tag_name, display_name, source_type, 
                       db_number, offset, data_type, bit_position,
                       string_length, byte_swap, unit, scaling, 
                       decimal_places, formula, mapping_name,
                       description, is_active
                FROM tags
                ORDER BY tag_name
            """)
            
            tags = cursor.fetchall()
            
            result = []
            for tag in tags:
                tag_dict = dict(tag)
                # Format PLC address
                if tag_dict.get('db_number') is not None and tag_dict.get('offset') is not None:
                    if tag_dict.get('bit_position') is not None:
                        tag_dict['plc_address'] = f"DB{tag_dict['db_number']}.{tag_dict['offset']}.{tag_dict['bit_position']}"
                    else:
                        tag_dict['plc_address'] = f"DB{tag_dict['db_number']}.{tag_dict['offset']}"
                result.append(tag_dict)
            
            return jsonify({
                'status': 'success',
                'tags': result,
                'count': len(result)
            })
    
    except Exception as e:
        logger.error(f"Error exporting tags: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': str(e)}), 500


@tags_bp.route('/tags/import-plc-csv', methods=['POST'])
def import_plc_csv():
    """Import tags from PLC engineering CSV files (exported from TIA Portal / Step 7).

    Accepts one or more CSV files. Each file has:
    - Row 1, Cell A1: the DB number (e.g. "2099")
    - Remaining rows: Name, DataType, Offset, Default, [flags...], Unit
    Struct rows are skipped (they are grouping containers).
    For Bool tags the offset column may contain "178.1" meaning byte 178 bit 1.
    """
    try:
        files = request.files.getlist('files')
        if not files:
            return jsonify({'status': 'error', 'message': 'No files provided'}), 400
        if len(files) > 20:
            return jsonify({'status': 'error', 'message': 'Too many files (max 20)'}), 400

        all_tags = []
        file_errors = []

        for f in files:
            try:
                content = f.read().decode('utf-8-sig')  # handle BOM
                if len(content) > 10 * 1024 * 1024:  # 10MB limit
                    file_errors.append({'file': f.filename, 'error': 'File too large (max 10MB)'})
                    continue
                reader = csv.reader(io.StringIO(content))
                rows = list(reader)

                if not rows or not rows[0]:
                    file_errors.append({'file': f.filename, 'error': 'Empty file'})
                    continue

                # Row 1: DB number is in the first cell
                db_number_raw = rows[0][0].strip()
                try:
                    db_number = int(db_number_raw)
                except ValueError:
                    file_errors.append({'file': f.filename, 'error': f'First cell is not a valid DB number: "{db_number_raw}"'})
                    continue

                # Process data rows (skip row 0 which is the DB number header)
                for row_idx, row in enumerate(rows[1:], start=2):
                    if not row or not row[0].strip():
                        continue

                    tag_name = row[0].strip()
                    data_type_raw = row[1].strip() if len(row) > 1 else ''

                    # Skip Struct rows - they are grouping containers
                    if data_type_raw.lower() == 'struct':
                        continue

                    offset_raw = row[2].strip() if len(row) > 2 else ''
                    if not offset_raw:
                        continue

                    # Map PLC data types to system types
                    type_map = {
                        'real': 'REAL',
                        'bool': 'BOOL',
                        'dint': 'DINT',
                        'int': 'INT',
                        'string': 'STRING',
                    }
                    data_type = type_map.get(data_type_raw.lower(), 'REAL')

                    # Parse offset - Bool can have "178.1" format (byte.bit)
                    bit_position = None
                    if '.' in offset_raw:
                        parts = offset_raw.split('.')
                        try:
                            offset = int(parts[0])
                            bit_position = int(parts[1])
                        except (ValueError, IndexError):
                            file_errors.append({'file': f.filename, 'error': f'Row {row_idx}: invalid offset "{offset_raw}"'})
                            continue
                    else:
                        try:
                            offset = int(offset_raw)
                        except ValueError:
                            file_errors.append({'file': f.filename, 'error': f'Row {row_idx}: invalid offset "{offset_raw}"'})
                            continue

                    # Unit is in the last non-empty column
                    unit = ''
                    if len(row) > 9 and row[9].strip():
                        unit = row[9].strip()
                    elif len(row) > 1:
                        # Fallback: last non-empty cell
                        for cell in reversed(row):
                            cell = cell.strip()
                            if cell and cell.upper() not in ('TRUE', 'FALSE', '0', '', 'REAL', 'INT', 'DINT', 'BOOL', 'STRING', 'STRUCT', 'WORD'):
                                try:
                                    float(cell)
                                except ValueError:
                                    unit = cell
                                    break

                    # Normalize unit abbreviations
                    unit_map = {
                        'ton per hour': 't/h',
                        'percentage per second': '%/s',
                        'leter per hour': 'L/h',
                        'litre per hour': 'L/h',
                        'liter per hour': 'L/h',
                    }
                    unit = unit_map.get(unit.lower(), unit)

                    # Build PLC address
                    if bit_position is not None:
                        plc_address = f"DB{db_number}.{offset}.{bit_position}"
                    else:
                        plc_address = f"DB{db_number}.{offset}"

                    # Clean tag name: replace spaces with underscores
                    clean_name = _sanitize_tag_name(tag_name)

                    all_tags.append({
                        'tag_name': clean_name,
                        'display_name': tag_name,
                        'source_type': 'PLC',
                        'plc_address': plc_address,
                        'data_type': data_type,
                        'unit': unit,
                        'is_active': True,
                    })

            except Exception as e:
                file_errors.append({'file': f.filename, 'error': str(e)})

        # Deduplicate: prefix with DB number, and offset if still not unique
        name_counts = {}
        for t in all_tags:
            name_counts[t['tag_name']] = name_counts.get(t['tag_name'], 0) + 1
        dupes = {n for n, c in name_counts.items() if c > 1}
        if dupes:
            for t in all_tags:
                if t['tag_name'] in dupes:
                    parsed_addr = parse_plc_address(t['plc_address'])
                    t['tag_name'] = f"DB{parsed_addr['db_number']}_{t['tag_name']}"
            # Check again for same-DB duplicates, append offset
            name_counts2 = {}
            for t in all_tags:
                name_counts2[t['tag_name']] = name_counts2.get(t['tag_name'], 0) + 1
            dupes2 = {n for n, c in name_counts2.items() if c > 1}
            if dupes2:
                for t in all_tags:
                    if t['tag_name'] in dupes2:
                        parsed_addr = parse_plc_address(t['plc_address'])
                        t['tag_name'] = f"{t['tag_name']}_off{parsed_addr['offset']}"

        if not all_tags:
            return jsonify({
                'status': 'error',
                'message': 'No valid tags found in the uploaded files',
                'file_errors': file_errors
            }), 400

        # Use existing bulk import logic
        imported = 0
        tag_errors = []

        get_db_connection_func = _get_db_connection()
        with closing(get_db_connection_func()) as conn:
            actual_conn = conn._conn if hasattr(conn, '_conn') else conn
            cursor = actual_conn.cursor()

            for tag_data in all_tags:
                try:
                    tag_name = tag_data['tag_name']
                    parsed = parse_plc_address(tag_data['plc_address'])
                    db_num = parsed['db_number']
                    off = parsed['offset']
                    bit_pos = parsed.get('bit')

                    cursor.execute("SELECT id FROM tags WHERE tag_name = %s", (tag_name,))
                    if cursor.fetchone():
                        cursor.execute("""
                            UPDATE tags SET
                                display_name = %s, source_type = %s, db_number = %s,
                                "offset" = %s, data_type = %s, bit_position = %s,
                                unit = %s, is_active = %s
                            WHERE tag_name = %s
                        """, (
                            tag_data.get('display_name', tag_name),
                            'PLC', db_num, off,
                            tag_data.get('data_type', 'REAL'),
                            bit_pos,
                            tag_data.get('unit', ''),
                            True,
                            tag_name
                        ))
                    else:
                        cursor.execute("""
                            INSERT INTO tags (
                                tag_name, display_name, source_type, db_number, "offset",
                                data_type, bit_position, unit, scaling, decimal_places,
                                description, is_active
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            tag_name,
                            tag_data.get('display_name', tag_name),
                            'PLC', db_num, off,
                            tag_data.get('data_type', 'REAL'),
                            bit_pos,
                            tag_data.get('unit', ''),
                            1.0, 2, '', True
                        ))

                    imported += 1
                except Exception as e:
                    tag_errors.append({'tag': tag_data.get('tag_name', '?'), 'error': str(e)})

            actual_conn.commit()

        # Auto-register in emulator if demo mode
        try:
            from demo_mode import get_demo_mode
            if get_demo_mode():
                from plc_data_source import register_tag_in_emulator
                for tag_data in all_tags:
                    try:
                        parsed = parse_plc_address(tag_data['plc_address'])
                        register_tag_in_emulator(
                            tag_data['tag_name'], 'PLC',
                            parsed['db_number'], parsed['offset'],
                            tag_data.get('data_type', 'REAL'),
                            tag_data.get('unit', '')
                        )
                    except Exception:
                        pass
        except Exception:
            pass

        return jsonify({
            'status': 'success',
            'imported': imported,
            'total_parsed': len(all_tags),
            'errors': tag_errors,
            'file_errors': file_errors,
            'message': f'Imported {imported} tags from {len(files)} file(s)'
        })

    except Exception as e:
        logger.error(f"Error importing PLC CSV: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': str(e)}), 500


@tags_bp.route('/tags/export-csv', methods=['GET'])
def export_tags_csv():
    """Export all tags as CSV file"""
    try:
        get_db_connection_func = _get_db_connection()
        with closing(get_db_connection_func()) as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            cursor.execute("""
                SELECT tag_name, display_name, source_type,
                       db_number, "offset", data_type, bit_position,
                       unit, scaling, decimal_places, formula,
                       mapping_name, description, is_active
                FROM tags
                ORDER BY tag_name
            """)

            tags = cursor.fetchall()

            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow([
                'tag_name', 'display_name', 'source_type', 'data_type',
                'plc_address', 'unit', 'scaling', 'decimal_places',
                'formula', 'mapping_name', 'description', 'is_active'
            ])

            for tag in tags:
                t = dict(tag)
                plc_address = ''
                if t.get('db_number') is not None and t.get('offset') is not None:
                    if t.get('bit_position') is not None:
                        plc_address = f"DB{t['db_number']}.{t['offset']}.{t['bit_position']}"
                    else:
                        plc_address = f"DB{t['db_number']}.{t['offset']}"

                writer.writerow([
                    t.get('tag_name', ''),
                    t.get('display_name', ''),
                    t.get('source_type', ''),
                    t.get('data_type', ''),
                    plc_address,
                    t.get('unit', ''),
                    t.get('scaling', 1.0),
                    t.get('decimal_places', 2),
                    t.get('formula', ''),
                    t.get('mapping_name', ''),
                    t.get('description', ''),
                    t.get('is_active', True),
                ])

            csv_content = output.getvalue()
            return jsonify({
                'status': 'success',
                'csv': csv_content,
                'count': len(tags)
            })

    except Exception as e:
        logger.error(f"Error exporting tags CSV: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': str(e)}), 500


@tags_bp.route('/tags/import-plc-excel', methods=['POST'])
def import_plc_excel():
    """Import tags from PLC engineering Excel files (.xlsx).

    Supports two formats:

    Format A (engineer spec sheets):
      - DB number extracted from sheet name (e.g. "Mill B DB499" -> 499)
      - Row 1: section title (skipped)
      - Row 2: column headers Name|Type|Offset|Comment (skipped)
      - Data rows: Name, Type, Offset, Comment (unit extracted from Comment)
      - Multiple sections per sheet separated by empty rows + new title/header

    Format B (TIA Portal CSV-style saved as .xlsx):
      - Cell A1 contains the DB number
      - Remaining rows: Name, Type, Offset, Default, [flags...], Unit (last col)
      - Struct rows skipped
    """
    try:
        try:
            import openpyxl
        except ImportError:
            return jsonify({'status': 'error', 'message': 'openpyxl library not installed — Excel import is unavailable'}), 500

        files = request.files.getlist('files')
        if not files:
            return jsonify({'status': 'error', 'message': 'No files provided'}), 400
        if len(files) > 20:
            return jsonify({'status': 'error', 'message': 'Too many files (max 20)'}), 400

        all_tags = []
        file_errors = []

        # Known unit patterns to extract from comment text
        unit_patterns = [
            't/h', 'kg', 'kg/h', '%', 'l/h', 'l/m', 'bar', 'rpm', 'mm/s',
            'kw', 'kwh', 'mw', 'hz', 'm³/h', 'm3/h', '°c', '°f',
        ]

        def extract_unit_from_comment(comment):
            """Try to extract a unit from comment text."""
            if not comment:
                return ''
            comment_lower = comment.strip().lower()
            # Direct match: comment IS the unit
            for u in unit_patterns:
                if comment_lower == u:
                    return comment.strip()
            # Check if comment starts with a known unit (e.g. "t/h" or "% Sender 1")
            for u in unit_patterns:
                if comment_lower.startswith(u + ' ') or comment_lower.startswith(u + ',') or comment_lower == u:
                    return u.upper() if len(u) <= 2 else u
            # Check for unit at end in parens: "Some desc (t/h)"
            paren_match = re.search(r'\(([^)]+)\)\s*$', comment)
            if paren_match:
                inner = paren_match.group(1).strip().lower()
                for u in unit_patterns:
                    if inner == u:
                        return paren_match.group(1).strip()
            return ''

        def detect_format(rows):
            """Detect if sheet is Format A (engineer spec) or Format B (TIA export)."""
            if not rows or not rows[0]:
                return 'A'
            a1 = str(rows[0][0] or '').strip()
            # Format B: first cell is just a number (DB number)
            try:
                int(a1)
                return 'B'
            except ValueError:
                pass
            return 'A'

        def extract_db_from_sheet_name(name):
            """Extract DB number from sheet name like 'Mill B DB499'."""
            match = re.search(r'DB\s*(\d+)', name, re.IGNORECASE)
            if match:
                return int(match.group(1))
            return None

        type_map = {
            'real': 'REAL', 'bool': 'BOOL', 'dint': 'DINT',
            'int': 'INT', 'string': 'STRING', 'word': 'INT',
        }

        unit_normalize = {
            'ton per hour': 't/h', 'percentage per second': '%/s',
            'leter per hour': 'L/h', 'litre per hour': 'L/h',
            'liter per hour': 'L/h',
        }

        for f in files:
            try:
                wb = openpyxl.load_workbook(f, data_only=True, read_only=True)

                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        continue

                    fmt = detect_format(rows)

                    if fmt == 'B':
                        # Format B: TIA-style, cell A1 = DB number
                        try:
                            db_number = int(str(rows[0][0]).strip())
                        except (ValueError, TypeError):
                            file_errors.append({'file': f.filename, 'sheet': sheet_name, 'error': 'Cell A1 is not a valid DB number'})
                            continue

                        for row_idx, row in enumerate(rows[1:], start=2):
                            if not row or not row[0]:
                                continue
                            tag_name = str(row[0]).strip()
                            data_type_raw = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                            if data_type_raw.lower() == 'struct':
                                continue
                            offset_raw = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
                            if not offset_raw:
                                continue

                            data_type = type_map.get(data_type_raw.lower(), 'REAL')
                            bit_position = None
                            if '.' in offset_raw:
                                parts = offset_raw.split('.')
                                try:
                                    offset = int(parts[0])
                                    bit_position = int(parts[1])
                                except (ValueError, IndexError):
                                    continue
                            else:
                                try:
                                    offset = int(float(offset_raw))
                                except ValueError:
                                    continue

                            # Unit from last non-empty, non-flag column
                            unit = ''
                            if len(row) > 9 and row[9]:
                                unit = str(row[9]).strip()
                            elif len(row) > 1:
                                for cell in reversed(row):
                                    if cell is None:
                                        continue
                                    cell_s = str(cell).strip()
                                    if cell_s and cell_s.upper() not in ('TRUE', 'FALSE', '0', ''):
                                        try:
                                            float(cell_s)
                                        except ValueError:
                                            unit = cell_s
                                            break

                            unit = unit_normalize.get(unit.lower(), unit)
                            plc_address = f"DB{db_number}.{offset}" if bit_position is None else f"DB{db_number}.{offset}.{bit_position}"
                            clean_name = _sanitize_tag_name(tag_name)

                            all_tags.append({
                                'tag_name': clean_name, 'display_name': tag_name,
                                'source_type': 'PLC', 'plc_address': plc_address,
                                'data_type': data_type, 'unit': unit, 'is_active': True,
                            })

                    else:
                        # Format A: engineer spec sheet, DB from sheet name
                        db_number = extract_db_from_sheet_name(sheet_name)
                        if db_number is None:
                            file_errors.append({'file': f.filename, 'sheet': sheet_name, 'error': f'Cannot extract DB number from sheet name "{sheet_name}"'})
                            continue

                        # Parse rows, skipping title rows and headers
                        for row_idx, row in enumerate(rows, start=1):
                            if not row or all(c is None for c in row):
                                continue

                            col_a = str(row[0]).strip() if row[0] is not None else ''
                            col_b = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
                            col_c = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
                            col_d = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''

                            # Skip header rows (Name|Type|Offset|Comment)
                            if col_a.lower() == 'name' and col_b.lower() == 'type':
                                continue

                            # Skip section title rows (only col_a filled, no type/offset)
                            if col_a and not col_b and not col_c:
                                continue

                            # Must have a valid type
                            data_type_raw = col_b.lower()
                            if data_type_raw not in type_map:
                                continue

                            data_type = type_map[data_type_raw]

                            # Parse offset
                            if not col_c:
                                continue
                            bit_position = None
                            if '.' in col_c:
                                parts = col_c.split('.')
                                try:
                                    offset = int(parts[0])
                                    bit_position = int(parts[1])
                                except (ValueError, IndexError):
                                    continue
                            else:
                                try:
                                    offset = int(float(col_c))
                                except ValueError:
                                    continue

                            # Extract unit from comment
                            unit = extract_unit_from_comment(col_d)

                            plc_address = f"DB{db_number}.{offset}" if bit_position is None else f"DB{db_number}.{offset}.{bit_position}"
                            clean_name = _sanitize_tag_name(col_a)

                            all_tags.append({
                                'tag_name': clean_name, 'display_name': col_a,
                                'source_type': 'PLC', 'plc_address': plc_address,
                                'data_type': data_type, 'unit': unit,
                                'description': col_d if col_d and col_d != unit else '',
                                'is_active': True,
                            })

            except Exception as e:
                file_errors.append({'file': f.filename, 'error': str(e)})
            finally:
                try:
                    wb.close()
                except Exception:
                    pass

        # Deduplicate: prefix with DB number, and offset if still not unique
        name_counts = {}
        for t in all_tags:
            name_counts[t['tag_name']] = name_counts.get(t['tag_name'], 0) + 1
        dupes = {n for n, c in name_counts.items() if c > 1}
        if dupes:
            for t in all_tags:
                if t['tag_name'] in dupes:
                    parsed_addr = parse_plc_address(t['plc_address'])
                    t['tag_name'] = f"DB{parsed_addr['db_number']}_{t['tag_name']}"
            # Check again for same-DB duplicates, append offset
            name_counts2 = {}
            for t in all_tags:
                name_counts2[t['tag_name']] = name_counts2.get(t['tag_name'], 0) + 1
            dupes2 = {n for n, c in name_counts2.items() if c > 1}
            if dupes2:
                for t in all_tags:
                    if t['tag_name'] in dupes2:
                        parsed_addr = parse_plc_address(t['plc_address'])
                        t['tag_name'] = f"{t['tag_name']}_off{parsed_addr['offset']}"

        if not all_tags:
            return jsonify({
                'status': 'error',
                'message': 'No valid tags found in the uploaded files',
                'file_errors': file_errors
            }), 400

        # Upsert into database
        imported = 0
        tag_errors = []

        get_db_connection_func = _get_db_connection()
        with closing(get_db_connection_func()) as conn:
            actual_conn = conn._conn if hasattr(conn, '_conn') else conn
            cursor = actual_conn.cursor()

            for tag_data in all_tags:
                try:
                    tag_name = tag_data['tag_name']
                    parsed = parse_plc_address(tag_data['plc_address'])
                    db_num = parsed['db_number']
                    off = parsed['offset']
                    bit_pos = parsed.get('bit')

                    cursor.execute("SELECT id FROM tags WHERE tag_name = %s", (tag_name,))
                    if cursor.fetchone():
                        cursor.execute("""
                            UPDATE tags SET
                                display_name = %s, source_type = %s, db_number = %s,
                                "offset" = %s, data_type = %s, bit_position = %s,
                                unit = %s, description = %s, is_active = %s
                            WHERE tag_name = %s
                        """, (
                            tag_data.get('display_name', tag_name),
                            'PLC', db_num, off, tag_data.get('data_type', 'REAL'),
                            bit_pos, tag_data.get('unit', ''),
                            tag_data.get('description', ''), True, tag_name
                        ))
                    else:
                        cursor.execute("""
                            INSERT INTO tags (
                                tag_name, display_name, source_type, db_number, "offset",
                                data_type, bit_position, unit, scaling, decimal_places,
                                description, is_active
                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            tag_name, tag_data.get('display_name', tag_name),
                            'PLC', db_num, off, tag_data.get('data_type', 'REAL'),
                            bit_pos, tag_data.get('unit', ''), 1.0, 2,
                            tag_data.get('description', ''), True
                        ))

                    imported += 1
                except Exception as e:
                    tag_errors.append({'tag': tag_data.get('tag_name', '?'), 'error': str(e)})

            actual_conn.commit()

        # Auto-register in emulator if demo mode
        try:
            from demo_mode import get_demo_mode
            if get_demo_mode():
                from plc_data_source import register_tag_in_emulator
                for tag_data in all_tags:
                    try:
                        parsed = parse_plc_address(tag_data['plc_address'])
                        register_tag_in_emulator(
                            tag_data['tag_name'], 'PLC',
                            parsed['db_number'], parsed['offset'],
                            tag_data.get('data_type', 'REAL'), tag_data.get('unit', '')
                        )
                    except Exception:
                        pass
        except Exception:
            pass

        return jsonify({
            'status': 'success',
            'imported': imported,
            'total_parsed': len(all_tags),
            'errors': tag_errors,
            'file_errors': file_errors,
            'message': f'Imported {imported} tags from {len(files)} file(s)'
        })

    except Exception as e:
        logger.error(f"Error importing PLC Excel: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': str(e)}), 500


@tags_bp.route('/tags/generate-report-drafts', methods=['POST'])
def generate_report_drafts():
    """Generate report builder template drafts from tags, grouped by DB number.

    Accepts optional JSON body:
      { "db_numbers": [199, 499, 899] }  — only generate for these DBs
    If omitted, generates for all DBs that have tags.

    Each report draft gets:
      - A table widget with all tags from that DB as columns
      - KPI cards for REAL-type tags (flow rates, percentages, etc.)
    """
    import uuid

    try:
        data = request.get_json(silent=True) or {}
        requested_dbs = data.get('db_numbers')
        sheet_names = data.get('sheet_names', {})  # {db_number: "sheet name"}

        # Validate db_numbers are integers
        if requested_dbs:
            try:
                requested_dbs = [int(d) for d in requested_dbs]
            except (ValueError, TypeError):
                return jsonify({'status': 'error', 'message': 'db_numbers must be a list of integers'}), 400

        get_db_connection_func = _get_db_connection()
        with closing(get_db_connection_func()) as conn:
            cursor = conn.cursor(cursor_factory=RealDictCursor)

            query = """
                SELECT tag_name, display_name, db_number, "offset", data_type,
                       bit_position, unit
                FROM tags
                WHERE source_type = 'PLC' AND is_active = TRUE
                  AND db_number IS NOT NULL
            """
            params = []
            if requested_dbs:
                placeholders = ','.join(['%s'] * len(requested_dbs))
                query += f" AND db_number IN ({placeholders})"
                params = requested_dbs

            query += " ORDER BY db_number, \"offset\""
            cursor.execute(query, params)
            tags = cursor.fetchall()

        if not tags:
            return jsonify({'status': 'error', 'message': 'No PLC tags found'}), 400

        # Group by db_number
        groups = {}
        for tag in tags:
            db = tag['db_number']
            if db not in groups:
                groups[db] = []
            groups[db].append(dict(tag))

        templates = []

        for db_number, db_tags in groups.items():
            report_name = sheet_names.get(str(db_number), sheet_names.get(db_number, f"DB{db_number} Report"))

            # Build paginatedSections for a tabular report
            sections = []

            # ── Header section ──
            sections.append({
                'id': f'ps-{uuid.uuid4().hex[:8]}',
                'type': 'header',
                'title': report_name,
                'subtitle': f'DB{db_number} — {len(db_tags)} tags',
                'showDateRange': True,
                'showLogo': False,
                'logoUrl': '',
                'align': 'center',
                'statusLabel': 'Status',
                'statusSourceType': 'static',
                'statusValue': '',
                'statusTagName': '',
                'statusMappingName': '',
                'statusFormula': '',
                'statusGroupTags': [],
                'statusAggregation': 'avg',
            })

            # ── Table section: tags as ROWS ──
            # Check if tags are counters/totalizers (DInt/Int types typically are)
            # Counter reports: Ident | Name | Total [unit] | Start [unit] | End [unit]
            # Regular reports: Ident | Name | Value [unit] | Unit
            has_counters = any(t['data_type'] in ('DINT', 'INT') for t in db_tags)
            has_reals = any(t['data_type'] == 'REAL' for t in db_tags)

            # If mix of types, use the general 5-column layout (works for both)
            if has_counters:
                # Counter/totalizer layout: show delta, start, end values
                # Determine a common unit label for headers
                units_in_group = set(t.get('unit', '') or '' for t in db_tags)
                common_unit = units_in_group.pop() if len(units_in_group) == 1 else ''
                total_hdr = f'Total {common_unit}'.strip() if common_unit else 'Total'
                start_hdr = f'StartValue {common_unit}'.strip() if common_unit else 'StartValue'
                end_hdr   = f'EndValue {common_unit}'.strip() if common_unit else 'EndValue'

                columns = [
                    {'id': f'ps-{uuid.uuid4().hex[:8]}', 'header': 'Ident',   'width': 'auto', 'align': 'left'},
                    {'id': f'ps-{uuid.uuid4().hex[:8]}', 'header': 'Name',    'width': 'auto', 'align': 'left'},
                    {'id': f'ps-{uuid.uuid4().hex[:8]}', 'header': total_hdr, 'width': 'auto', 'align': 'right'},
                    {'id': f'ps-{uuid.uuid4().hex[:8]}', 'header': start_hdr, 'width': 'auto', 'align': 'right'},
                    {'id': f'ps-{uuid.uuid4().hex[:8]}', 'header': end_hdr,   'width': 'auto', 'align': 'right'},
                ]

                rows = []
                for tag in db_tags:
                    unit = tag.get('unit', '') or ''
                    decimals = 2 if tag['data_type'] == 'REAL' else 0
                    display = tag.get('display_name') or tag['tag_name']
                    cell_unit = '' if common_unit else unit  # avoid double unit display

                    rows.append({
                        'id': f'ps-{uuid.uuid4().hex[:8]}',
                        'cells': [
                            {'sourceType': 'static', 'value': tag['tag_name']},
                            {'sourceType': 'static', 'value': display},
                            # Total = delta (end - start)
                            {'sourceType': 'tag', 'tagName': tag['tag_name'], 'value': '',
                             'formula': '', 'unit': cell_unit, 'decimals': decimals,
                             'groupTags': [], 'aggregation': 'delta', 'mappingName': ''},
                            # Start = first value in period
                            {'sourceType': 'tag', 'tagName': tag['tag_name'], 'value': '',
                             'formula': '', 'unit': cell_unit, 'decimals': decimals,
                             'groupTags': [], 'aggregation': 'first', 'mappingName': ''},
                            # End = last value in period
                            {'sourceType': 'tag', 'tagName': tag['tag_name'], 'value': '',
                             'formula': '', 'unit': cell_unit, 'decimals': decimals,
                             'groupTags': [], 'aggregation': 'last', 'mappingName': ''},
                        ],
                    })
            else:
                # Regular (REAL/BOOL) layout: Ident | Name | Value | Unit
                columns = [
                    {'id': f'ps-{uuid.uuid4().hex[:8]}', 'header': 'Ident',  'width': 'auto', 'align': 'left'},
                    {'id': f'ps-{uuid.uuid4().hex[:8]}', 'header': 'Name',   'width': 'auto', 'align': 'left'},
                    {'id': f'ps-{uuid.uuid4().hex[:8]}', 'header': 'Value',  'width': 'auto', 'align': 'right'},
                    {'id': f'ps-{uuid.uuid4().hex[:8]}', 'header': 'Unit',   'width': 'auto', 'align': 'center'},
                ]

                rows = []
                for tag in db_tags:
                    unit = tag.get('unit', '') or ''
                    decimals = 2 if tag['data_type'] == 'REAL' else 0
                    display = tag.get('display_name') or tag['tag_name']

                    rows.append({
                        'id': f'ps-{uuid.uuid4().hex[:8]}',
                        'cells': [
                            {'sourceType': 'static', 'value': tag['tag_name']},
                            {'sourceType': 'static', 'value': display},
                            {'sourceType': 'tag', 'tagName': tag['tag_name'], 'value': '',
                             'formula': '', 'unit': '', 'decimals': decimals,
                             'groupTags': [], 'aggregation': 'last', 'mappingName': ''},
                            {'sourceType': 'static', 'value': unit},
                        ],
                    })

            sections.append({
                'id': f'ps-{uuid.uuid4().hex[:8]}',
                'type': 'table',
                'label': f'DB{db_number} Tags',
                'columns': columns,
                'rows': rows,
                'showSummaryRow': False,
                'summaryLabel': 'Total',
                'summaryFormula': '',
                'summaryUnit': '',
            })

            template = {
                'id': uuid.uuid4().hex,
                'name': report_name,
                'description': f'Auto-generated draft from DB{db_number} import ({len(db_tags)} tags)',
                'status': 'draft',
                'layout_config': {
                    'reportType': 'paginated',
                    'schemaVersion': 2,
                    'paginatedSections': sections,
                    'widgets': [],
                    'parameters': [],
                    'computedSignals': [],
                    'grid': {
                        'cols': 12,
                        'rowHeight': 40,
                        'pageMode': 'a4',
                    },
                },
                'created_at': None,
                'updated_at': None,
            }
            templates.append(template)

        return jsonify({
            'status': 'success',
            'templates': templates,
            'count': len(templates),
            'message': f'Generated {len(templates)} report draft(s)',
        })

    except Exception as e:
        logger.error(f"Error generating report drafts: {e}", exc_info=True)
        return jsonify({'status': 'error', 'message': str(e)}), 500

